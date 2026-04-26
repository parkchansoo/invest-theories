#!/usr/bin/env python3
"""
extract_trades.py — Parse all broker PDF files and extract trade data.

Outputs:
  - trades_all.csv (date, broker, type, stock, quantity, amount, fees)
  - trades_by_stock.xlsx (one sheet per stock with all trades)
  - deposits.json (real external deposits/withdrawals)
  - trades_summary.md (yearly summary table)
  - Comparison report vs heatmap ground truth

Usage: python3 extract_trades.py
"""
import os, re, json, csv
from collections import defaultdict
from datetime import datetime

import pdfplumber

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.dirname(SCRIPT_DIR)
RAW_DIR = os.path.join(DATA_DIR, "raw_data", "originals")
PARSED_DIR = os.path.join(DATA_DIR, "raw_data", "parsed")
OUTPUT_DIR = os.path.join(DATA_DIR, "output")
GROUND_TRUTH = os.path.join(OUTPUT_DIR, "trade_data.json")

# ============================================================
# STOCK NAME MAPPING
# ============================================================
STOCK_NAME_MAP = {
    # 키움 English → ground truth Korean
    "APPLE": "애플", "애플": "애플",
    "TESLA": "테슬라", "TESLA INC": "테슬라", "테슬라": "테슬라",
    "ALPHABET INC": "알파벳 A", "ALPHABET INC-CL A": "알파벳 A",
    "ALPHABET INC-CL C": "알파벳 C", "알파벳 A": "알파벳 A", "알파벳 C": "알파벳 C",
    "AMAZON.COM": "아마존 닷컴", "AMAZON.COM INC": "아마존 닷컴", "아마존 닷컴": "아마존 닷컴", "아마존": "아마존 닷컴",
    "NVIDIA CORP": "엔비디아", "NVIDIA": "엔비디아", "엔비디아": "엔비디아",
    "MICROSOFT CORP": "마이크로소프트", "마이크로소프트": "마이크로소프트",
    "STARBUCKS CORP": "스타벅스", "스타벅스": "스타벅스",
    "BERKSHIRE HATHAWAY INC-CL A": "버크셔 해서웨이 A", "버크셔 해서웨이 A": "버크셔 해서웨이 A",
    "REALTY INCOME CORP": "리얼티 인컴", "리얼티 인컴": "리얼티 인컴",
    "NIKE INC -CL B": "나이키", "나이키 B": "나이키", "나이키": "나이키",
    "AT&T INC": "AT&T", "AT&T": "AT&T",
    "HERSHEY CO/THE": "허쉬", "허쉬": "허쉬",
    "PEPSICO INC": "펩시코", "펩시코": "펩시코",
    "TAIWAN SEMICONDUCTOR-SP ADR": "TSMC ADR",
    "CIRCLE INTERNET GROUP INC": "써클 인터넷 그룹", "써클 인터넷 그룹": "써클 인터넷 그룹",
    "ROBINHOOD MARKETS INC": "로빈훗 마케츠", "로빈후드": "로빈훗 마케츠", "로빈훗 마케츠": "로빈훗 마케츠",
    "FACEBOOK INC-CLASS A": "페이스북", "META PLATFORMS INC-CLASS A": "페이스북",
    "BITDEER TECHNOLOGIES GROUP": "비트마인 이머션 테크놀로지스", "비트마인 이머션 테크놀로지스": "비트마인 이머션 테크놀로지스",
    "ALBEMARLE CORP": "앨버말", "앨버말": "앨버말",
    "COCA-COLA CO/THE": "코카콜라", "코카콜라": "코카콜라",
    "ALTRIA GROUP INC": "알트리아 그룹", "알트리아 그룹": "알트리아 그룹",
    "PALANTIR TECHNOLOGIES INC": "팔란티어", "팔란티어": "팔란티어",
    "WARNER BROS DISCOVERY INC": "WARNER BROS DISCOVERY INC",
    "PLUG POWER INC": "플러그 파워", "플러그 파워": "플러그 파워",
    "ASML HOLDING NV-NY": "ASML 홀딩",
    "ENDRA LIFE SCIENCES INC": "엔드라 라이프 사이언시스", "엔드라 라이프 사이언시스": "엔드라 라이프 사이언시스",
    "ONEOK INC": "원오케이", "원오케이": "원오케이",
    "ARK AUTONOMOUS TECHNOLOGY &": "ARKQ (ARK Autonomous Technolog",
    "RESOLVE AI INC": "리졸브 AI", "리졸브 AI": "리졸브 AI",
    "COUPANG INC": "쿠팡", "쿠팡": "쿠팡",
    "VELOCITYSHARES 3X LONG CRUDE": "UWT (VelocityShares 3x Long Cr",
    "FIRST REPUBLIC BANK/CA": "퍼스트리퍼블릭", "퍼스트리퍼블릭": "퍼스트리퍼블릭",
    "BROADCOM INC": "브로드컴", "브로드컴": "브로드컴",
    "LASER PHOTONICS CORP": "레이저 포토닉스", "레이저 포토닉스": "레이저 포토닉스",
    "PROKIDNEY CORP": "프로키드니", "프로키드니": "프로키드니",
    "IMPACT BIOMEDICAL INC": "임팩트 바이오메디컬", "임팩트 바이오메디컬": "임팩트 바이오메디컬",
    "MAIN STREET CAPITAL CORP": "메인 스트리트 캐피탈", "메인 스트리트 캐피탈": "메인 스트리트 캐피탈",
    "PETROLEO BRASILEIRO-SPON ADR": "페트로브라스 우선주", "페트로브라스 우선주(ADR)": "페트로브라스 우선주", "페트로브라스 우선주": "페트로브라스 우선주",
    "GREENLAND ACQUISITION CORP": "그린랜드 마인즈", "그린랜드 마인즈": "그린랜드 마인즈",
    "BUZZFEED INC": "버즈피드", "버즈피드": "버즈피드",
    "MOONLAKE IMMUNOTHERAPEUTICS": "문레이크 이뮤노테라퓨틱스", "문레이크 이뮤노테라퓨틱스": "문레이크 이뮤노테라퓨틱스",
    "LONGEVERON INC": "롱에버론", "롱에버론": "롱에버론",
    "RALLY BIO CORP": "랠리바이오", "랠리바이오": "랠리바이오",
    "ASPIRE BIOPHARMA HOLDINGS IN": "어스파이어 바이오파머 홀딩스", "어스파이어 바이오파머 홀딩스": "어스파이어 바이오파머 홀딩스",
    "CBDMD INC": "CBDMD", "CBDMD": "CBDMD",
    "STEAKHOLDER FOODS LTD": "스테이크홀더 푸즈", "스테이크홀더 푸즈": "스테이크홀더 푸즈",
    "BED BATH & BEYOND": "베드 배스 앤 비욘드", "베드 배스 앤 비욘드": "베드 배스 앤 비욘드",
    "BIT DIGITAL INC": "비트 디지털", "비트 디지털": "비트 디지털",
    "ARDELYX INC": "아델릭스", "아델릭스": "아델릭스",
    "GORILLA TECHNOLOGY GROUP INC": "고릴라 테크놀로지 그룹", "고릴라 테크놀로지 그룹": "고릴라 테크놀로지 그룹",
    "KAIXIN AUTO HOLDINGS": "캉고", "캉고": "캉고",
    "ADAPTIMMUNE THERAPEUTICS PLC": "어댑트이뮨 테라퓨틱스", "어댑트이뮨 테라퓨틱스": "어댑트이뮨 테라퓨틱스",
    "BRIGHT GREEN CORP": "브라이트 그린", "브라이트 그린": "브라이트 그린",
    "CANOO INC": "카누", "카누": "카누",
    "JOHNSON & JOHNSON": "존슨 앤 존슨", "존슨 앤 존슨": "존슨 앤 존슨",
    "PROGENITY INC": "프로팬크 바이오파마", "프로팬크 바이오파마": "프로팬크 바이오파마",
    # ETFs
    "YIELDMAX NVDA OPTION INCOME": "NVDY (일드맥스 엔비디아 옵션배당 전략 ETF)", "일드맥스 엔비디아 옵션배당 전략 ETF": "NVDY (일드맥스 엔비디아 옵션배당 전략 ETF)",
    "SPDR DJ INDUSTRIAL AVERAGE ETF": "DIA (SPDR DJ Industrial Averag", "SPDR DJ INDUSTRIAL AVERAGE": "DIA (SPDR DJ Industrial Averag", "SPDR 다우존스 ETF": "DIA (SPDR DJ Industrial Averag",
    "SPDR S&P 500": "SPY (SPDR S&P 500)",
    "YIELDMAX TSLA OPTION INCOME": "TSLY (일드맥스 테슬라 옵션 인컴 전략 ETF)", "일드맥스 테슬라 옵션 인컴 전략 ETF": "TSLY (일드맥스 테슬라 옵션 인컴 전략 ETF)",
    "VANGUARD S&P 500 ETF": "VOO (Vanguard S&P 500 ETF)", "뱅가드 S&P500 ETF": "VOO (Vanguard S&P 500 ETF)",
    "JPMORGAN NASDAQ EQUITY PREMIU": "JEPQ (JP모건 나스닥 프리미엄 인컴 ETF)", "JP모건 나스닥 프리미엄 인컴 ETF": "JEPQ (JP모건 나스닥 프리미엄 인컴 ETF)",
    "SCHWAB US DIVIDEND EQUITY ETF": "SCHD (슈왑 미국 배당주 ETF)", "슈왑 미국 배당주 ETF": "SCHD (슈왑 미국 배당주 ETF)",
    "GLOBAL X LITHIUM & BATTERY TECH ETF": "LIT (Global X Lithium & Batter", "GLOBAL X LITHIUM & BATTERY": "LIT (Global X Lithium & Batter", "글로벌엑스 리튬 배터리 ETF": "LIT (Global X Lithium & Batter",
    "PROETF ULTRAPRO QQQ": "TQQQ (프로셰어즈 QQQ 3배 ETF)", "PROSHARES ULTRAPRO QQQ": "TQQQ (프로셰어즈 QQQ 3배 ETF)", "프로셰어즈 QQQ 3배 ETF": "TQQQ (프로셰어즈 QQQ 3배 ETF)",
    "INVESCO QQQ TRUST": "QQQ (Invesco QQQ Trust)",
    "YIELDMAX COIN OPTION INCOME": "CONY (일드맥스 코인베이스 옵션 배당 ETF)", "일드맥스 코인베이스 옵션 배당 ETF": "CONY (일드맥스 코인베이스 옵션 배당 ETF)",
    "ARK INNOVATION ETF": "ARKK (ARK Innovation ETF)",
    "JPMORGAN EQUITY PREMIUM INCOME ETF": "JEPI (JPMorgan Equity Premium ", "JP모건 커버드콜 옵션 ETF": "JEPI (JPMorgan Equity Premium ",
    "SPDR S&P 500 GROWTH ETF": "SPYG (SPDR S&P 500 Growth ETF)", "SPDR S&P500 포트폴리오 ETF": "SPYG (SPDR S&P 500 Growth ETF)",
    "WISDOMTREE US QUALITY DIVID": "DGRW (위즈덤트리 퀄리티 배당 성장 ETF)", "위즈덤트리 퀄리티 배당 성장 ETF": "DGRW (위즈덤트리 퀄리티 배당 성장 ETF)",
    "SHELL MIDSTREAM PARTNERS LP": "SHLX (Shell Midstream Partners",
    "ISHARES IBOXX HIGH YIELD COR": "HYG (블랙록 고수익 회사채 ETF)", "블랙록 고수익 회사채 ETF": "HYG (블랙록 고수익 회사채 ETF)",
    "SKILLZ INC": "SKLZ (Skillz Inc)",
    "ISHARES MSCI INDIA ETF": "아이셰어즈 인도 ETF", "아이셰어즈 인도 ETF": "아이셰어즈 인도 ETF",
    "DIREXION DAILY FTSE CHINA BULL 3X ETF": "YINN (Direxion Daily FTSE Chin", "DIREXION DAILY FTSE CHINA BULL": "YINN (Direxion Daily FTSE Chin",
    "VANECK SOCIAL SENTIMENT ETF": "BUZZ (VanEck Social Sentiment ",
    "VANECK VECTORS SOCIAL SENTIMENT ETF": "BUZZ (VanEck Social Sentiment ",
    "DIREXION DAILY SEMICOND BULL": "SOXL (디렉시온 미국 반도체 3배 ETF)", "디렉시온 미국 반도체 3배 ETF": "SOXL (디렉시온 미국 반도체 3배 ETF)",
    "SACHEM CAPITAL CORP": "SACH (Sachem Capital)",
    "ISHARES MSCI SOUTH KOREA ETF": "EWY (아이셰어즈 대한민국 ETF)", "아이셰어즈 대한민국 ETF": "EWY (아이셰어즈 대한민국 ETF)",
    "DIREXION DAILY 20+ YEAR TREA": "TMF (디렉시온 20년 미국채 3배 ETF)", "디렉시온 20년 미국채 3배 ETF": "TMF (디렉시온 20년 미국채 3배 ETF)",
    "GUGGENHEIM TAXABLE MUNI BOND": "구겐하임 지방채 펀드",
    "PROETF ULTRAPRO SHORT QQQ": "SQQQ (ProShares UltraPro Short", "PROSHARES ULTRAPRO SHORT QQQ": "SQQQ (ProShares UltraPro Short",
    # Korean domestic
    "삼성전자": "삼성전자", "삼성전자우": "삼성전자우", "삼성SDI": "삼성SDI", "삼성전기": "삼성전기",
    "카카오": "카카오", "LG생활건강": "LG생활건강", "LG전자": "LG전자", "현대차": "현대차",
    "NAVER": "NAVER", "하이브": "하이브", "빅히트": "하이브",
    "피씨엘": "피씨엘", "원익IPS": "원익IPS", "하나금융지주": "하나금융지주",
    "태경케미컬": "태경케미컬", "서울옥션": "서울옥션", "일진디스플": "일진디스플",
    "한국전력": "한국전력", "비츠로셀": "비츠로셀", "현대제철": "현대제철", "대유": "대유",
    "한화손해보험": "한화손해보험", "신일전자": "신일전자", "삼부토건": "삼부토건",
    "서울식품": "서울식품", "JYP Ent.": "JYP Ent.", "월덱스": "월덱스", "HMM": "HMM",
    "휴마시스": "휴마시스", "대한전선": "대한전선", "와이씨": "와이씨",
    "케이엔알시스템": "케이엔알시스템", "데이타솔루션": "데이타솔루션",
    "초록뱀미디어": "초록뱀미디어", "세종텔레콤": "세종텔레콤", "우리종합금융": "우리종합금융",
    "고려신용정보": "고려신용정보", "비바리퍼블리카": "비바리퍼블리카", "대유 7R": "대유 7R",
    "대유 7R(J2903801C)": "대유 7R",
    # 토스 additional Korean names
    "아마존닷컴": "아마존 닷컴",
    "TSMC": "TSMC ADR",
    "워너 브로스 디스커버리": "WARNER BROS DISCOVERY INC",
    "쉘 미드스트림 파트너스": "SHLX (Shell Midstream Partners",
    "구겐하임 지방채": "구겐하임 지방채 펀드",
    "SPDR S&P 500 포트폴리오 ETF": "SPYG (SPDR S&P 500 Growth ETF)",
    # 키움 multi-line ETF names (as they appear after joining)
    "VELOCITYSHARES 3X LONG CRUDE OIL ETN": "UWT (VelocityShares 3x Long Cr",
    "DIREXION DAILY FTSE CHINA BULL 3X": "YINN (Direxion Daily FTSE Chin",
    "JPMORGAN EQUITY PREMIUM INCOME": "JEPI (JPMorgan Equity Premium ",
    "JPMORGAN NASDAQ EQUITY PREMIUM INCOME": "JEPQ (JP모건 나스닥 프리미엄 인컴 ETF)",
    "JPMORGAN EQUITY PREMIUM INCOME ETF": "JEPI (JPMorgan Equity Premium ",
    # 연금저축
    "TIGER 미국나스닥100": "TIGER 미국나스닥100", "ACE 미국S&P500": "ACE 미국S&P500",
    "SOL 미국배당다우존스": "SOL 미국배당다우존스", "SOL미국배당다우존스": "SOL 미국배당다우존스",
    "TIGER 인도니프티50": "TIGER 인도니프티50", "TIGER 배당성장": "TIGER 배당성장", "TIGER배당성장": "TIGER 배당성장",
    "ACE 미국배당퀄리티": "ACE 미국배당퀄리티", "ACE미국배당퀄리티": "ACE 미국배당퀄리티",
    "TIGER 미국테크TOP10타겟커버드콜": "TIGER 미국테크TOP10타겟커버드콜",
    "KODEX 미국S&P500TR": "KODEX 미국S&P500TR",
    "TIGER 배당커버드콜액티브": "TIGER 배당커버드콜액티브", "TIGER배당커버드콜액티브": "TIGER 배당커버드콜액티브",
    "TIGER 미국S&P500": "TIGER 미국S&P500",
}

# Stock splits: (stock_name, split_date, split_ratio)
STOCK_SPLITS = [
    ("애플", "2020-08-28", 4),
    ("테슬라", "2020-08-28", 5),
    ("테슬라", "2022-08-25", 3),
    ("알파벳 A", "2022-07-15", 20),
    ("알파벳 C", "2022-07-15", 20),
    ("아마존 닷컴", "2022-06-06", 20),
    ("엔비디아", "2021-07-20", 4),
    ("엔비디아", "2024-06-11", 10),  # Split effective June 10 open, but Toss records pre-split buy on 6/10
]


def pn(s):
    """Parse number, removing commas."""
    if not s or s == '-': return 0
    return float(s.strip().replace(',', ''))


def normalize_stock(name):
    """Map PDF stock name to ground truth name."""
    name = name.strip()
    # Remove various code suffixes
    name = re.sub(r'\(A\d+\)$', '', name).strip()
    name = re.sub(r'\(US[A-Z0-9]+\)$', '', name).strip()
    name = re.sub(r'\([A-Z]{2}[A-Z0-9]+\)$', '', name).strip()
    name = re.sub(r'\(J\d+[A-Z]+\)$', '', name).strip()  # 대유 7R code
    if name in STOCK_NAME_MAP: return STOCK_NAME_MAP[name]
    up = name.upper()
    for k, v in STOCK_NAME_MAP.items():
        if k.upper() == up: return v
    # Partial match (longer prefix)
    best_match = None; best_len = 0
    for k, v in STOCK_NAME_MAP.items():
        ku = k.upper()
        if len(ku) > 5:
            # Check if either starts with the other
            common = min(len(ku), len(up))
            match_len = 0
            for ci in range(common):
                if ku[ci] == up[ci]: match_len += 1
                else: break
            if match_len > best_len and match_len >= min(8, len(ku)):
                best_len = match_len
                best_match = v
    if best_match: return best_match
    return name


def get_split_factor(stock_name, trade_date):
    """Get the cumulative split factor for a stock as of the trade date."""
    factor = 1
    for sname, sdate, sratio in STOCK_SPLITS:
        if stock_name == sname and trade_date < sdate:
            factor *= sratio
    return factor


# ============================================================
# KIWOOM PARSER
# ============================================================
def parse_kiwoom(pdf_path):
    trades, deposits, exrates = [], [], {}
    pdf = pdfplumber.open(pdf_path)

    for page in pdf.pages:
        text = page.extract_text()
        if not text: continue
        lines = text.split('\n')
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            # Match: YYYY/MM/DD ...rest...
            m = re.match(r'^(\d{4}/\d{2}/\d{2})\s+(.+)$', line)
            if not m:
                i += 1; continue

            date_str = m.group(1)
            rest = m.group(2)

            # Determine where the currency line is:
            # Pattern A (normal): date+stock+nums / currency+type
            # Pattern B (multi-line): [prev line: stock prefix] / date+nums / [next: stock suffix] / currency+type
            # Pattern C (multi-line): date+stock_part+nums / [next: stock suffix] / currency+type

            # Check if line i+1 is currency
            nxt1 = lines[i+1].strip() if i+1 < len(lines) else ''
            nxt2 = lines[i+2].strip() if i+2 < len(lines) else ''
            nxt3 = lines[i+3].strip() if i+3 < len(lines) else ''

            cm1 = re.match(r'^(KRW|USD)\s+(.+)$', nxt1)
            cm2 = re.match(r'^(KRW|USD)\s+(.+)$', nxt2)
            cm3 = re.match(r'^(KRW|USD)\s+(.+)$', nxt3)

            stock_prefix = ''  # from line before date
            stock_suffix = ''  # from line after date (before currency)

            if cm1:
                # Pattern A: normal 2-line record
                # But check if previous line was a stock name fragment
                if i > 0:
                    prev = lines[i-1].strip()
                    # If prev line is not a date line, not a currency line, not a header/noise
                    if (prev and not re.match(r'^\d{4}/', prev) and not re.match(r'^(KRW|USD)\s', prev)
                        and '거래일자' not in prev and '잔고' not in prev and '계좌' not in prev
                        and '수량' not in prev and '■' not in prev and '※' not in prev
                        and '2026-03-20' not in prev and not prev.startswith('*')
                        and not re.match(r'^page', prev, re.IGNORECASE)):
                        # Check if rest looks like it's missing a stock name (starts with pure numbers)
                        parts1 = rest.split()
                        potential_stock = ' '.join(parts1[:-5]) if len(parts1) >= 5 else ''
                        if not potential_stock or all(re.match(r'^[\d,.]+$', p) for p in potential_stock.split()):
                            stock_prefix = prev
                cm = cm1
                skip_lines = 2
            elif cm2:
                # Pattern C: stock suffix on nxt1, currency on nxt2
                stock_suffix = nxt1
                cm = cm2
                skip_lines = 3
                # Also check for prefix from previous line
                if i > 0:
                    prev = lines[i-1].strip()
                    if (prev and not re.match(r'^\d{4}/', prev) and not re.match(r'^(KRW|USD)\s', prev)
                        and '거래일자' not in prev and '잔고' not in prev and '계좌' not in prev
                        and '수량' not in prev and '■' not in prev and '※' not in prev
                        and '2026-03-20' not in prev and not prev.startswith('*')
                        and not re.match(r'^page', prev, re.IGNORECASE)):
                        stock_prefix = prev
            elif cm3:
                # Even more complex: prefix + date + suffix + currency
                if i > 0:
                    prev = lines[i-1].strip()
                    if prev and not re.match(r'^\d{4}/', prev) and not re.match(r'^(KRW|USD)\s', prev):
                        stock_prefix = prev
                stock_suffix = nxt1
                # nxt2 might also be suffix or might be something else
                if not re.match(r'^(KRW|USD)\s', nxt2):
                    stock_suffix = nxt1 + ' ' + nxt2
                cm = cm3
                skip_lines = 4
            else:
                i += 1; continue

            # Build full stock+numbers line
            full_rest = rest
            if stock_prefix:
                # Prepend stock prefix to rest
                parts1 = rest.split()
                if len(parts1) >= 5:
                    potential_stock = ' '.join(parts1[:-5])
                    if not potential_stock or all(re.match(r'^[\d,.]+$', p) for p in potential_stock.split() if p):
                        full_rest = stock_prefix + ' ' + rest
            if stock_suffix:
                # Insert stock suffix: it goes between any existing stock name and the numeric fields
                parts1 = full_rest.split()
                if len(parts1) >= 5:
                    potential_stock = ' '.join(parts1[:-5])
                    nums = parts1[-5:]
                    full_rest = potential_stock + ' ' + stock_suffix + ' ' + ' '.join(nums)

            currency = cm.group(1)
            type_rest = cm.group(2)

            # Parse line 1: stock qty amount fees penalty balance
            fr_parts = full_rest.split()
            # Parse from right: balance, penalty, fees, amount, qty → stock is the rest
            if len(fr_parts) < 5:
                i += skip_lines; continue
            try:
                balance1 = pn(fr_parts[-1])
                penalty1 = pn(fr_parts[-2])
                fees1 = pn(fr_parts[-3])
                amount1 = pn(fr_parts[-4])
                qty1 = pn(fr_parts[-5])
                stock = ' '.join(fr_parts[:-5])
            except: i += skip_lines; continue

            # Parse line 2: type price foreign_amount tax income_tax foreign_balance
            parts2 = type_rest.split()
            if len(parts2) < 6:
                i += skip_lines; continue
            try:
                tx_type = parts2[0]
                price2 = pn(parts2[1])
                famount2 = pn(parts2[2])
                tax2 = pn(parts2[3])
                itax2 = pn(parts2[4])
                fbal2 = pn(parts2[5])
            except: i += skip_lines; continue

            ndate = date_str.replace('/', '-')

            # Exchange rate extraction
            if '원화대용자동외화매수' in tx_type:
                if price2 > 0 and famount2 > 0:
                    exrates[ndate] = price2
                i += skip_lines; continue

            # Skip noise
            skip = ['원화대용해지입금','원화대용지정출금','원화대용지정출금(지급결제)',
                    '원화대용자동외화매수(휴일)','원화대용자동외화매수',
                    '예탁금이용료(이자)입금','환전정산입금','업무수수료출금',
                    '해외이벤트입금','배당금(외화)입금','배당금입금',
                    '해외배당금입금']
            if tx_type in skip: i += skip_lines; continue

            # Deposits
            if tx_type in ['이체입금(지급결제)','이체입금(은행간)','이체입금','연금저축정기납입']:
                if currency == 'KRW' and amount1 > 0:
                    deposits.append({'date': ndate, 'type': 'deposit', 'amount': int(amount1), 'broker': '키움', 'description': tx_type})
                i += skip_lines; continue

            if '이체출금' in tx_type and '대체' not in tx_type and 'TOSS' not in tx_type.upper():
                if currency == 'KRW' and amount1 > 0:
                    if '토스증권' not in tx_type and '키움증권' not in tx_type:
                        deposits.append({'date': ndate, 'type': 'withdrawal', 'amount': int(amount1), 'broker': '키움', 'description': tx_type})
                i += skip_lines; continue

            if '대체TOSS' in tx_type or ('대체' in tx_type and '출금' in tx_type):
                i += skip_lines; continue

            # Trades
            if tx_type == '매수':
                if qty1 < 1:
                    i += skip_lines; continue
                ns = normalize_stock(stock)
                if currency == 'USD' and famount2 > 0 and stock:
                    trades.append({'date': ndate, 'broker': '키움', 'type': 'buy', 'stock': ns,
                                   'quantity': int(qty1) if qty1 == int(qty1) else qty1,
                                   'amount_usd': famount2, 'amount': 0, 'fees': fees1, 'currency': 'USD'})
                elif currency == 'KRW' and amount1 > 0 and stock:
                    trades.append({'date': ndate, 'broker': '키움', 'type': 'buy', 'stock': ns,
                                   'quantity': int(qty1), 'amount': int(amount1), 'fees': fees1, 'currency': 'KRW'})
                i += skip_lines; continue

            if tx_type == '매도':
                if qty1 < 1:
                    i += skip_lines; continue
                ns = normalize_stock(stock)
                if currency == 'USD' and famount2 > 0 and stock:
                    trades.append({'date': ndate, 'broker': '키움', 'type': 'sell', 'stock': ns,
                                   'quantity': int(qty1) if qty1 == int(qty1) else qty1,
                                   'amount_usd': famount2, 'amount': 0, 'fees': fees1, 'tax': tax2, 'currency': 'USD'})
                elif currency == 'KRW' and amount1 > 0 and stock:
                    trades.append({'date': ndate, 'broker': '키움', 'type': 'sell', 'stock': ns,
                                   'quantity': int(qty1), 'amount': int(amount1), 'fees': fees1, 'tax': tax2, 'currency': 'KRW'})
                i += skip_lines; continue

            if '타사대체출고' in tx_type and stock and qty1 > 0:
                trades.append({'date': ndate, 'broker': '키움', 'type': 'transfer_out', 'stock': normalize_stock(stock),
                               'quantity': int(qty1), 'amount': 0, 'fees': 0, 'currency': currency})
                i += skip_lines; continue

            if '타사대체입고' in tx_type and stock and qty1 > 0:
                trades.append({'date': ndate, 'broker': '키움', 'type': 'transfer_in', 'stock': normalize_stock(stock),
                               'quantity': int(qty1), 'amount': 0, 'fees': 0, 'currency': currency})
                i += skip_lines; continue

            i += skip_lines
    pdf.close()
    return trades, deposits, exrates


# ============================================================
# TOSS PARSER
# ============================================================
def parse_toss(pdf_path):
    trades, deposits = [], []
    pdf = pdfplumber.open(pdf_path)
    in_dollar = False

    for page in pdf.pages:
        text = page.extract_text()
        if not text: continue
        lines = text.split('\n')

        for i, line in enumerate(lines):
            line = line.strip()
            if '달러 거래내역' in line: in_dollar = True; continue
            if '원화 거래내역' in line: in_dollar = False; continue

            m = re.match(r'^(\d{4}\.\d{2}\.\d{2})\s+(.+)$', line)
            if not m: continue
            dstr = m.group(1).replace('.', '-')
            rest = m.group(2)

            if in_dollar:
                _parse_toss_usd(dstr, rest, lines, i, trades)
            else:
                _parse_toss_krw(dstr, rest, trades, deposits)
    pdf.close()
    return trades, deposits


def _parse_toss_krw(dstr, rest, trades, deposits):
    parts = rest.split()
    if len(parts) < 3: return
    tx = parts[0]

    # Deposits/withdrawals
    if tx.startswith(('이체입금', '이체출금', '오픈뱅킹입금')):
        # Format: type name 0 amount 0 0 0 0 0 0 balance
        # Find amount: look for the first large number after the name
        nums = re.findall(r'[\d,]+', rest)
        for n in nums:
            v = pn(n)
            if v >= 1:  # Skip the "0" placeholders
                is_dep = '입금' in tx
                is_broker = any(x in tx for x in ['토스증권', '키움증권'])
                if not is_broker:
                    deposits.append({'date': dstr, 'type': 'deposit' if is_dep else 'withdrawal',
                                     'amount': int(v), 'broker': '토스', 'description': tx})
                break
        return

    # Stock trades
    if tx in ['구매', '판매']:
        # Parse from right: 9 numeric fields (balance2, balance, penalty, deduction, tax, fee, price, amount, qty)
        # Stock name is between tx and the numeric fields
        all_tokens = parts[1:]  # Everything after tx
        nums_from_right = []
        stock_end = len(all_tokens)
        for j in range(len(all_tokens)-1, -1, -1):
            try:
                pn(all_tokens[j])
                # Check it looks numeric (might have commas)
                if re.match(r'^[\d,]+$', all_tokens[j]) or all_tokens[j] == '0':
                    nums_from_right.insert(0, pn(all_tokens[j]))
                    stock_end = j
                else:
                    break
            except:
                break

        stock_name = ' '.join(all_tokens[:stock_end])
        if len(nums_from_right) >= 9:
            qty = nums_from_right[0]
            amount = nums_from_right[1]
            price = nums_from_right[2]
            fee = nums_from_right[3]
            tax = nums_from_right[4]

            if stock_name and amount > 0:
                trades.append({
                    'date': dstr, 'broker': '토스',
                    'type': 'buy' if tx == '구매' else 'sell',
                    'stock': normalize_stock(stock_name),
                    'quantity': int(qty) if qty == int(qty) else qty,
                    'amount': int(amount), 'fees': fee, 'tax': tax, 'currency': 'KRW'
                })
        return

    # Transfers
    if '타사대체입고' in tx or '대체입고' in tx:
        cm = re.search(r',\s*(.+?)(?:\(A\d+\))?\s+(\d+)\s+', rest)
        if cm:
            trades.append({'date': dstr, 'broker': '토스', 'type': 'transfer_in',
                           'stock': normalize_stock(cm.group(1).strip()),
                           'quantity': int(cm.group(2)), 'amount': 0, 'fees': 0, 'currency': 'KRW'})
        return

    if '대체출고' in tx:
        cm = re.search(r',\s*(.+?)(?:\(A\d+\))?\s+(\d+)\s+', rest)
        if cm:
            trades.append({'date': dstr, 'broker': '토스', 'type': 'transfer_out',
                           'stock': normalize_stock(cm.group(1).strip()),
                           'quantity': int(cm.group(2)), 'amount': 0, 'fees': 0, 'currency': 'KRW'})
        return

    if '대행예탁' in tx:
        m2 = re.search(r'(\S+?)(?:\(A\d+\))?\s+(\d[\d,]*)\s+', rest)
        if m2:
            qty = int(pn(m2.group(2)))
            if qty > 0:
                trades.append({'date': dstr, 'broker': '토스', 'type': 'transfer_in',
                               'stock': normalize_stock(m2.group(1)), 'quantity': qty,
                               'amount': 0, 'fees': 0, 'currency': 'KRW'})
        return

    if '친구초대이벤트입고' in tx:
        cm = re.search(r',\s*(.+?)(?:\(A\d+\))?\s+(\d+)\s+', rest)
        if cm:
            trades.append({'date': dstr, 'broker': '토스', 'type': 'transfer_in',
                           'stock': normalize_stock(cm.group(1).strip()),
                           'quantity': int(cm.group(2)), 'amount': 0, 'fees': 0, 'currency': 'KRW'})
        return


def _parse_toss_usd(dstr, rest, lines, line_idx, trades):
    parts = rest.split()
    if len(parts) < 3: return
    tx = parts[0]

    if tx not in ['구매', '판매']: return

    # Parse from right for numeric fields
    all_tokens = parts[1:]
    nums_from_right = []
    stock_end = len(all_tokens)
    for j in range(len(all_tokens)-1, -1, -1):
        try:
            val = all_tokens[j].replace(',', '')
            float(val)
            nums_from_right.insert(0, pn(all_tokens[j]))
            stock_end = j
        except:
            break

    stock_name = ' '.join(all_tokens[:stock_end])
    # Remove code suffix
    stock_name = re.sub(r'\(US[A-Z0-9]+\)$', '', stock_name).strip()
    stock_name = re.sub(r'\([A-Z]{2}[A-Z0-9]+\)$', '', stock_name).strip()

    # Dollar section has: rate qty krw_amount price fee tax penalty balance balance2
    # That's 9 numeric fields
    if len(nums_from_right) >= 9:
        rate = nums_from_right[0]
        qty = nums_from_right[1]
        krw_amount = nums_from_right[2]
        price = nums_from_right[3]
        fee = nums_from_right[4]
        tax = nums_from_right[5]

        if stock_name and krw_amount > 0:
            trades.append({
                'date': dstr, 'broker': '토스',
                'type': 'buy' if tx == '구매' else 'sell',
                'stock': normalize_stock(stock_name),
                'quantity': int(qty) if qty == int(qty) else qty,
                'amount': int(krw_amount), 'fees': fee, 'tax': tax,
                'currency': 'USD', 'exchange_rate': rate
            })


# ============================================================
# SHINHAN PARSER
# ============================================================
def parse_shinhan(pdf_path):
    trades, deposits = [], []
    pdf = pdfplumber.open(pdf_path)

    for page in pdf.pages:
        text = page.extract_text()
        if not text: continue
        lines = text.split('\n')
        i = 0
        while i < len(lines):
            m = re.match(r'^(\d{4}-\d{2}-\d{2})\s+(.*)$', lines[i].strip())
            if not m: i += 1; continue
            dstr = m.group(1)
            rest1 = m.group(2)
            if i + 2 >= len(lines): i += 1; continue
            line2 = lines[i+1].strip()
            line3 = lines[i+2].strip()
            parts2 = line2.split()
            if len(parts2) < 2: i += 1; continue
            try: seq = int(parts2[0]); tx = parts2[1]
            except: i += 1; continue

            if tx in ['장내_매수', '장내매수']:
                p1 = rest1.split()
                stock = ''; price = 0; fee = 0
                nums1 = []; sp1 = []
                for j in range(len(p1)-1, -1, -1):
                    try: float(p1[j].replace(',','')); nums1.insert(0, p1[j])
                    except: sp1 = p1[:j+1]; break
                stock = ' '.join(sp1)
                if len(nums1) >= 2: price = pn(nums1[0]); fee = pn(nums1[1])
                qty = pn(parts2[2]) if len(parts2) >= 3 else 0
                total = int(price * qty) if price > 0 and qty > 0 else 0
                if stock and total > 0:
                    trades.append({'date': dstr, 'broker': '신한', 'type': 'buy',
                                   'stock': normalize_stock(stock), 'quantity': int(qty),
                                   'amount': total, 'fees': fee, 'currency': 'KRW'})
                i += 3; continue

            if tx in ['장내_매도', '장내매도']:
                p1 = rest1.split()
                stock = ''; price = 0; fee = 0
                nums1 = []; sp1 = []
                for j in range(len(p1)-1, -1, -1):
                    try: float(p1[j].replace(',','')); nums1.insert(0, p1[j])
                    except: sp1 = p1[:j+1]; break
                stock = ' '.join(sp1)
                if len(nums1) >= 2: price = pn(nums1[0]); fee = pn(nums1[1])
                qty = pn(parts2[2]) if len(parts2) >= 3 else 0
                tax = pn(parts2[3]) if len(parts2) >= 4 else 0
                total = int(price * qty) if price > 0 and qty > 0 else 0
                if stock and total > 0:
                    trades.append({'date': dstr, 'broker': '신한', 'type': 'sell',
                                   'stock': normalize_stock(stock), 'quantity': int(qty),
                                   'amount': total, 'fees': fee, 'tax': tax, 'currency': 'KRW'})
                i += 3; continue

            if tx == '전자이체입금':
                nums3 = [pn(p) for p in line3.split() if re.match(r'^[\d,]+$', p)]
                if len(nums3) >= 3 and nums3[2] > 0:
                    deposits.append({'date': dstr, 'type': 'deposit', 'amount': int(nums3[2]),
                                     'broker': '신한', 'description': tx})
                i += 3; continue

            if tx == '전자이체출금':
                nums3 = [pn(p) for p in line3.split() if re.match(r'^[\d,]+$', p)]
                if len(nums3) >= 3 and nums3[2] > 0:
                    deposits.append({'date': dstr, 'type': 'withdrawal', 'amount': int(nums3[2]),
                                     'broker': '신한', 'description': tx})
                i += 3; continue

            i += 3
    pdf.close()
    return trades, deposits


# ============================================================
# MIRAE ASSET PARSER
# ============================================================
def parse_mirae(pdf_path):
    trades, deposits = [], []
    pdf = pdfplumber.open(pdf_path)
    for pi, page in enumerate(pdf.pages):
        text = page.extract_text()
        if not text or ('증 명 서' in text and pi == 0): continue
        lines = text.split('\n')
        i = 0
        while i < len(lines):
            m = re.match(r'^(\d{4}/\d{2}/\d{2})(.*)$', lines[i].strip())
            if not m: i += 1; continue
            dstr = m.group(1).replace('/', '-')
            rest = m.group(2).strip()

            if '주식매수입고' in rest:
                stock = ''
                if i+1 < len(lines):
                    sn = lines[i+1].strip()
                    sn = re.sub(r'증권상장지수투자?신탁$', '', sn)
                    sn = re.sub(r'증권상장지수투자?$', '', sn)
                    sn = re.sub(r'증권상장$', '', sn)
                    sn = re.sub(r'^신한 ', '', sn)
                    sn = re.sub(r'^삼성 ', '', sn)
                    stock = sn.strip()
                qty = 0; price = 0
                if i+2 < len(lines):
                    p3 = lines[i+2].strip().split()
                    if len(p3) >= 4:
                        try: qty = pn(p3[2]); price = pn(p3[3])
                        except: pass
                total = int(price * qty) if price > 0 and qty > 0 else 0
                if stock and total > 0:
                    mapped = normalize_stock(stock)
                    if 'SOL' in stock and '미국배당' in stock: mapped = 'SOL 미국배당다우존스'
                    elif 'KODEX' in stock: mapped = 'KODEX 미국S&P500TR'
                    trades.append({'date': dstr, 'broker': '미래에셋', 'type': 'buy',
                                   'stock': mapped, 'quantity': int(qty),
                                   'amount': total, 'fees': 0, 'currency': 'KRW'})
                i += 3
                while i < len(lines) and '주식매수출금' in lines[i]: i += 2
                continue

            if '이체입금' in rest and '분배금' not in rest:
                nums = [pn(n) for n in re.findall(r'[\d,]+', rest.replace('이체입금', ''))]
                if nums:
                    amt = max(nums)
                    if amt > 0:
                        deposits.append({'date': dstr, 'type': 'deposit', 'amount': int(amt),
                                         'broker': '미래에셋', 'description': '이체입금'})
                i += 2; continue

            i += 1
    pdf.close()
    return trades, deposits


# ============================================================
# KAKAOPAY PARSER
# ============================================================
def parse_kakaopay(pdf_path):
    trades, deposits = [], []
    pdf = pdfplumber.open(pdf_path)
    for page in pdf.pages:
        text = page.extract_text()
        if not text: continue
        lines = text.split('\n')
        i = 0
        while i < len(lines):
            m = re.match(r'^(\d{4}\.\d{2}\.\d{2})\s+(.+)$', lines[i].strip())
            if not m: i += 1; continue
            dstr = m.group(1).replace('.', '-')
            rest = m.group(2)
            parts = rest.split()
            if len(parts) < 2: i += 1; continue
            tx = parts[0]

            if '매수입고' in tx:
                price = pn(parts[1]) if len(parts) > 1 else 0
                qty = pn(parts[2]) if len(parts) > 2 else 0
                stock = ''
                if i+2 < len(lines):
                    dm = re.match(r'^\d{2}:\d{2}:\d{2}\s+(.+?)\s+\d', lines[i+2].strip())
                    if dm: stock = dm.group(1).strip()
                total = int(price * qty) if price > 0 and qty > 0 else 0
                if stock and total > 0:
                    trades.append({'date': dstr, 'broker': '카카오페이', 'type': 'buy',
                                   'stock': normalize_stock(stock), 'quantity': int(qty),
                                   'amount': total, 'fees': 0, 'currency': 'KRW'})
                i += 3; continue

            if '매수출금' in tx:
                # Update fees on last buy
                if len(parts) >= 2:
                    settlement = pn(parts[-1])
                    if trades and trades[-1]['type'] == 'buy' and trades[-1]['date'] == dstr:
                        fee = settlement - trades[-1]['amount']
                        if fee > 0: trades[-1]['fees'] = fee
                i += 3; continue

            if '대체입금(연금저축정기납입)' in tx or '연금저축정기납입' in tx:
                nums = [pn(n) for n in re.findall(r'[\d,]+', rest)]
                if nums:
                    amt = max(nums)
                    if amt > 0:
                        deposits.append({'date': dstr, 'type': 'deposit', 'amount': int(amt),
                                         'broker': '카카오페이', 'description': '연금저축정기납입'})
                i += 3; continue

            if '전자망이체입금' in tx:
                nums = [pn(n) for n in re.findall(r'[\d,]+', rest)]
                if nums:
                    amt = max(nums)
                    if amt > 0:
                        deposits.append({'date': dstr, 'type': 'deposit', 'amount': int(amt),
                                         'broker': '카카오페이', 'description': '전자망이체입금'})
                i += 3; continue

            i += 1
    pdf.close()
    return trades, deposits


# ============================================================
# EXCHANGE RATE FOR KIWOOM USD TRADES
# ============================================================
def compute_kiwoom_krw(trades, exrates):
    sorted_rates = sorted(exrates.items())
    for t in trades:
        if t['broker'] == '키움' and t.get('currency') == 'USD' and t.get('amount_usd', 0) > 0:
            rate = _nearest_rate(t['date'], sorted_rates)
            if rate:
                t['amount'] = round(t['amount_usd'] * rate)
                t['exchange_rate'] = rate
            else:
                yr = int(t['date'][:4])
                fb = {2019:1170, 2020:1180, 2021:1150, 2022:1290, 2023:1300, 2024:1350}
                r = fb.get(yr, 1300)
                t['amount'] = round(t['amount_usd'] * r)
                t['exchange_rate'] = r


def _nearest_rate(target, sorted_rates):
    if not sorted_rates: return None
    best = None; bd = 999
    for ds, r in sorted_rates:
        try:
            d = abs((datetime.strptime(target, '%Y-%m-%d') - datetime.strptime(ds, '%Y-%m-%d')).days)
            if d < bd: bd = d; best = r
        except: pass
    return best


# ============================================================
# AGGREGATION WITH SPLIT ADJUSTMENT
# ============================================================
def aggregate(all_trades):
    cells = defaultdict(lambda: {'b':0,'sl':0,'bc':0,'sc':0,'tin':0,'tout':0})
    for t in all_trades:
        month = t['date'][:7]
        stock = t['stock']
        key = f"{month}|{stock}"
        sf = get_split_factor(stock, t['date'])
        qty = t['quantity']
        adjusted_qty = int(qty * sf) if isinstance(qty, (int, float)) and qty > 0 else 0

        if t['type'] == 'buy':
            cells[key]['b'] += t['amount']
            cells[key]['bc'] += adjusted_qty
        elif t['type'] == 'sell':
            cells[key]['sl'] += t['amount']
            cells[key]['sc'] += adjusted_qty
        elif t['type'] == 'transfer_in':
            cells[key]['tin'] += adjusted_qty
        elif t['type'] == 'transfer_out':
            cells[key]['tout'] += adjusted_qty
    return dict(cells)


def compare(computed, gt_path):
    with open(gt_path) as f: gt = json.load(f)
    gt_cells = gt['cells']
    allk = set(list(computed.keys()) + list(gt_cells.keys()))
    matches = mismatches = missing = extra = 0
    details = []
    for k in sorted(allk):
        gv = gt_cells.get(k); cv = computed.get(k)
        if gv and not cv:
            missing += 1; details.append(f"  MISSING: {k} → GT={gv}")
        elif cv and not gv:
            extra += 1
            if any(v != 0 for v in cv.values()):
                details.append(f"  EXTRA: {k} → COMP={cv}")
        elif gv == cv: matches += 1
        else:
            mismatches += 1
            details.append(f"  MISMATCH: {k}\n    GT:   {gv}\n    COMP: {cv}")

    total = matches + mismatches + missing
    rate = (matches / total * 100) if total > 0 else 0
    print(f"\n{'='*60}\nCOMPARISON REPORT\n{'='*60}")
    print(f"GT cells: {len(gt_cells)} | Computed: {len(computed)}")
    print(f"Match: {matches} | Mismatch: {mismatches} | Missing: {missing} | Extra: {extra}")
    print(f"Match rate: {rate:.1f}%\n")
    for d in details[:80]: print(d)
    return rate


# ============================================================
# OUTPUT
# ============================================================
def write_csv(trades, path):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['date','broker','type','stock','quantity','amount','fees','tax'])
        for t in sorted(trades, key=lambda x: x['date']):
            w.writerow([t['date'], t['broker'], t['type'], t['stock'],
                        t['quantity'], t['amount'], t.get('fees',0), t.get('tax',0)])
    print(f"Wrote {len(trades)} trades to {path}")


def write_deposits(deps, path):
    deps.sort(key=lambda x: x['date'])
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(deps, f, ensure_ascii=False, indent=2)
    print(f"Wrote {len(deps)} deposits to {path}")


def classify_stock(name):
    """종목을 국내/해외/국내 etf/해외 etf로 분류"""
    # 국내 ETF: ACE, KODEX, SOL, TIGER, 미래에셋 TIGER
    kr_etf_prefixes = ('ACE ', 'KODEX ', 'SOL ', 'TIGER ', '미래에셋 TIGER')
    if any(name.startswith(p) for p in kr_etf_prefixes):
        return '국내 etf'

    # 해외 ETF: 괄호 안에 ETF 설명이 있는 패턴 or 알려진 해외 ETF
    foreign_etf_names = {
        'ARKK', 'ARKQ', 'BUZZ', 'CONY', 'DGRW', 'DIA', 'EWY', 'HYG',
        'JEPI', 'JEPQ', 'LIT', 'NVDY', 'QQQ', 'SCHD', 'SHLX', 'SOXL',
        'SPY', 'SPYG', 'SQQQ', 'TMF', 'TQQQ', 'TSLY', 'UWT', 'VOO', 'YINN',
    }
    ticker = name.split(' (')[0].split('(')[0].strip()
    if ticker in foreign_etf_names:
        return '해외 etf'
    if '아이셰어즈' in name and 'ETF' in name:
        return '해외 etf'
    if '구겐하임' in name and '펀드' in name:
        return '해외 etf'

    # 국내 주식: 한글 이름 or 알려진 한국 종목
    kr_stocks = {
        'HMM', 'JYP Ent.', 'NAVER',
        'LG생활건강', 'LG전자', '삼성SDI', '삼성전기', '삼성전자', '삼성전자우',
        '고려신용정보', '대유', '대유 7R', '대한전선', '데이타솔루션',
        '비바리퍼블리카', '비츠로셀', '삼부토건', '서울식품', '서울옥션',
        '세종텔레콤', '신일전자', '와이씨', '우리종합금융', '원익IPS',
        '월덱스', '일진디스플', '초록뱀미디어', '카카오', '케이엔알시스템',
        '태경케미컬', '피씨엘', '하나금융지주', '하이브', '한국전력',
        '한화손해보험', '현대제철', '현대차', '휴마시스',
    }
    if name in kr_stocks:
        return '국내'

    # 나머지는 해외 주식
    return '해외'


def _get_splits_for_stock(stock):
    """해당 종목의 액면분할 이력 반환"""
    return [(sdate, sratio) for sname, sdate, sratio in STOCK_SPLITS if sname == stock]


def _remaining_split_factor(stock, trade_date):
    """거래일 이후 남은 분할 배수 (현재 기준 조정용)"""
    factor = 1
    for sname, sdate, sratio in STOCK_SPLITS:
        if sname == stock and trade_date < sdate:
            factor *= sratio
    return factor


def _fmt_krw(n):
    """원화 포맷: 1,234,567원"""
    if n == 0: return '0원'
    if isinstance(n, float) and n < 1:
        return f'{n:,.2f}원'
    return f'{int(round(n)):,}원'


def _fmt_usd(n):
    """달러 포맷: $123.45"""
    if n == 0: return '$0'
    if n >= 1:
        return f'${n:,.2f}'
    return f'${n:,.4f}'


def _fmt_qty(n):
    """수량 포맷: 정수면 정수, 소수면 유효숫자 2자리"""
    if isinstance(n, float) and n != int(n):
        return f'{n:.2g}'
    return f'{int(n)}'


def _per_share_prices(t):
    """거래에서 주당가격(원화, 달러) 계산. (krw, usd) 튜플 반환. 계산 불가 시 None."""
    qty = t['quantity']
    amount = t['amount']
    if qty == 0 or amount == 0:
        return None, None

    krw_per_share = amount / qty

    # USD 계산
    usd_per_share = None
    if t.get('amount_usd') and t['amount_usd'] > 0:
        usd_per_share = t['amount_usd'] / qty
    elif t.get('exchange_rate') and t['exchange_rate'] > 0:
        usd_per_share = krw_per_share / t['exchange_rate']

    return krw_per_share, usd_per_share


def write_stock_md(trades, base_dir):
    """종목별 마크다운 파일 생성 — output/stock/{카테고리}/{종목명}.md"""
    import shutil

    stock_dir = os.path.join(base_dir, 'stock')
    if os.path.exists(stock_dir):
        shutil.rmtree(stock_dir)

    categories = ['국내', '해외', '국내 etf', '해외 etf']
    for cat in categories:
        os.makedirs(os.path.join(stock_dir, cat), exist_ok=True)

    by_stock = defaultdict(list)
    for t in sorted(trades, key=lambda x: x['date']):
        by_stock[t['stock']].append(t)

    counts = defaultdict(int)
    for stock, stock_trades in sorted(by_stock.items()):
        cat = classify_stock(stock)
        counts[cat] += 1
        is_foreign = cat in ('해외', '해외 etf')

        safe_name = re.sub(r'[\\/*?:"<>|]', '', stock)
        filepath = os.path.join(stock_dir, cat, f'{safe_name}.md')

        # 통계
        buys = [t for t in stock_trades if t['type'] == 'buy']
        sells = [t for t in stock_trades if t['type'] == 'sell']
        transfers_in = [t for t in stock_trades if t['type'] == 'transfer_in']
        transfers_out = [t for t in stock_trades if t['type'] == 'transfer_out']

        total_buy_amt = sum(t['amount'] for t in buys)
        total_sell_amt = sum(t['amount'] for t in sells)
        total_buy_qty = sum(t['quantity'] for t in buys)
        total_sell_qty = sum(t['quantity'] for t in sells)
        total_fees = sum(t.get('fees', 0) for t in stock_trades)
        total_tax = sum(t.get('tax', 0) for t in stock_trades)
        brokers = sorted(set(t['broker'] for t in stock_trades))
        first_date = stock_trades[0]['date']
        last_date = stock_trades[-1]['date']

        splits = _get_splits_for_stock(stock)
        has_splits = len(splits) > 0

        lines = [f'# {stock}\n']
        lines.append('| 항목 | 값 |')
        lines.append('|------|-----|')
        lines.append(f'| 카테고리 | {cat} |')
        lines.append(f'| 증권사 | {", ".join(brokers)} |')
        lines.append(f'| 기간 | {first_date} ~ {last_date} |')
        lines.append(f'| 총 매수 | {_fmt_krw(total_buy_amt)} ({_fmt_qty(total_buy_qty)}주) |')
        lines.append(f'| 총 매도 | {_fmt_krw(total_sell_amt)} ({_fmt_qty(total_sell_qty)}주) |')
        if transfers_in:
            lines.append(f'| 대체입고 | {_fmt_qty(sum(t["quantity"] for t in transfers_in))}주 |')
        if transfers_out:
            lines.append(f'| 대체출고 | {_fmt_qty(sum(t["quantity"] for t in transfers_out))}주 |')
        lines.append(f'| 수수료 합계 | {_fmt_krw(total_fees)} |')
        if total_tax > 0:
            lines.append(f'| 세금 합계 | {_fmt_krw(total_tax)} |')
        lines.append('')

        # ── 액면분할 이력 ──
        if has_splits:
            lines.append('## 액면분할 이력\n')
            lines.append('| 날짜 | 비율 | 누적 배수 |')
            lines.append('|------|------|----------|')
            cum = 1
            for sdate, sratio in sorted(splits):
                cum *= sratio
                lines.append(f'| {sdate} | 1:{sratio} | 1:{cum} |')
            lines.append('')

        # ── 거래 내역 (대체입고/출고 제외) ──
        trade_rows = [t for t in stock_trades if t['type'] in ('buy', 'sell')]

        lines.append('## 거래 내역\n')
        if is_foreign:
            if has_splits:
                lines.append('| 날짜 | 증권사 | 유형 | 수량 | 총금액 | 주당(원) | 주당($) | 액분후 수량 | 액분후 주당(원) | 액분후 주당($) | 수수료 |')
                lines.append('|------|--------|------|-----:|-------:|--------:|-------:|----------:|-----------:|----------:|-------:|')
            else:
                lines.append('| 날짜 | 증권사 | 유형 | 수량 | 총금액 | 주당(원) | 주당($) | 수수료 |')
                lines.append('|------|--------|------|-----:|-------:|--------:|-------:|-------:|')
        else:
            if has_splits:
                lines.append('| 날짜 | 증권사 | 유형 | 수량 | 총금액 | 주당(원) | 액분후 수량 | 액분후 주당(원) | 수수료 |')
                lines.append('|------|--------|------|-----:|-------:|--------:|----------:|-----------:|-------:|')
            else:
                lines.append('| 날짜 | 증권사 | 유형 | 수량 | 총금액 | 주당(원) | 수수료 |')
                lines.append('|------|--------|------|-----:|-------:|--------:|-------:|')

        for t in trade_rows:
            type_kr = '매수' if t['type'] == 'buy' else '매도'
            qty = t['quantity']
            amount = t['amount']
            fees = t.get('fees', 0)

            krw_ps, usd_ps = _per_share_prices(t)
            krw_str = _fmt_krw(krw_ps) if krw_ps is not None else '-'
            usd_str = _fmt_usd(usd_ps) if usd_ps is not None else '-'

            # 분할 조정
            sf = _remaining_split_factor(stock, t['date'])
            if has_splits and sf > 1 and qty > 0:
                adj_qty = qty * sf
                adj_krw = _fmt_krw(krw_ps / sf) if krw_ps else '-'
                adj_usd = _fmt_usd(usd_ps / sf) if usd_ps else '-'
            elif has_splits:
                adj_qty = qty
                adj_krw = krw_str
                adj_usd = usd_str
            else:
                adj_qty = None

            if is_foreign and has_splits:
                row = f'| {t["date"]} | {t["broker"]} | {type_kr} | {_fmt_qty(qty)} | {_fmt_krw(amount)} | {krw_str} | {usd_str} | {_fmt_qty(adj_qty)} | {adj_krw} | {adj_usd} | {_fmt_krw(fees)} |'
            elif is_foreign:
                row = f'| {t["date"]} | {t["broker"]} | {type_kr} | {_fmt_qty(qty)} | {_fmt_krw(amount)} | {krw_str} | {usd_str} | {_fmt_krw(fees)} |'
            elif has_splits:
                row = f'| {t["date"]} | {t["broker"]} | {type_kr} | {_fmt_qty(qty)} | {_fmt_krw(amount)} | {krw_str} | {_fmt_qty(adj_qty)} | {adj_krw} | {_fmt_krw(fees)} |'
            else:
                row = f'| {t["date"]} | {t["broker"]} | {type_kr} | {_fmt_qty(qty)} | {_fmt_krw(amount)} | {krw_str} | {_fmt_krw(fees)} |'
            lines.append(row)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines) + '\n')

    summary = ', '.join(f'{cat} {counts[cat]}개' for cat in categories if counts[cat])
    print(f"Wrote {sum(counts.values())} stock MD files to {stock_dir} ({summary})")


def write_by_stock_xlsx(trades, path):
    """trades_all → 종목별 시트로 분리된 XLSX 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Group by stock
    by_stock = defaultdict(list)
    for t in sorted(trades, key=lambda x: x['date']):
        by_stock[t['stock']].append(t)

    wb = Workbook()
    wb.remove(wb.active)

    headers = ['date', 'broker', 'type', 'quantity', 'amount', 'fees', 'tax']
    header_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell_font = Font(name='맑은 고딕', size=10)
    cell_align = Alignment(vertical='center')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
    )
    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    for stock in sorted(by_stock.keys()):
        sheet_name = re.sub(r'[\\/*?\[\]:]', '', stock)[:31] or 'Sheet'
        # Deduplicate sheet names
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + '_' + str(len(existing))
        ws = wb.create_sheet(title=sheet_name)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = thin_border

        for ri, t in enumerate(by_stock[stock], 2):
            vals = [t['date'], t['broker'], t['type'], t['quantity'], t['amount'], t.get('fees', 0), t.get('tax', 0)]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
                if ri % 2 == 0:
                    cell.fill = alt_fill

        # Column widths
        widths = [12, 10, 14, 10, 16, 12]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(by_stock[stock]) + 1}'
        ws.freeze_panes = 'A2'

    wb.save(path)
    print(f"Wrote {len(by_stock)} stock sheets to {path}")


def _identify_inter_account_transfers(trades):
    """내계좌 간 대체입출고를 식별하여 해당 trade의 id set을 반환."""
    from collections import defaultdict
    by_date_stock = defaultdict(list)
    for t in trades:
        if t['type'] in ('transfer_in', 'transfer_out'):
            by_date_stock[(t['date'], t['stock'])].append(t)

    inter = set()
    # Pattern 1: 같은 날 같은 종목, 서로 다른 증권사에서 입고+출고 → 내계좌 대체
    for (date, stock), records in by_date_stock.items():
        brokers = set(r['broker'] for r in records)
        types = set(r['type'] for r in records)
        if len(brokers) > 1 and len(types) > 1:
            for r in records:
                inter.add(id(r))

    # Pattern 2: 키움 대량 출고 (2023-07-24 키움→토스) — 같은 날 단일 증권사에서 여러 종목 출고
    kiwoom_bulk_out_dates = set()
    for (date, stock), records in by_date_stock.items():
        if all(r['broker'] == '키움' and r['type'] == 'transfer_out' for r in records):
            kiwoom_bulk_out_dates.add(date)
    # 같은 날 5종목 이상 출고면 대체출고로 판단
    bulk_dates = set()
    for d in kiwoom_bulk_out_dates:
        count = sum(1 for (dd, _) in by_date_stock if dd == d)
        if count >= 5:
            bulk_dates.add(d)
    for (date, stock), records in by_date_stock.items():
        if date in bulk_dates:
            for r in records:
                inter.add(id(r))

    # Pattern 3: 키움 대량 입고 (2020-11-04 타사→키움) — 같은 날 단일 증권사 입고만
    kiwoom_bulk_in_dates = set()
    for (date, stock), records in by_date_stock.items():
        if all(r['broker'] == '키움' and r['type'] == 'transfer_in' for r in records):
            kiwoom_bulk_in_dates.add(date)
    for d in kiwoom_bulk_in_dates:
        count = sum(1 for (dd, _) in by_date_stock if dd == d
                    and all(r['broker'] == '키움' and r['type'] == 'transfer_in' for r in by_date_stock[(dd, _)]))
        # 이 날 키움 입고가 3건 이상이면 내계좌 대체
        in_count = sum(1 for (dd, s) in by_date_stock if dd == d)
        if in_count >= 3:
            for (dd, s), recs in by_date_stock.items():
                if dd == d and all(r['broker'] == '키움' and r['type'] == 'transfer_in' for r in recs):
                    for r in recs:
                        inter.add(id(r))

    return inter


def _parse_krw_cell(s):
    if not s or s.strip() == '': return None
    s = s.strip().replace('₩','').replace(',','')
    neg = s.startswith('-')
    s = s.replace('-','')
    v = int(s) if s else 0
    return -v if neg else v


def _load_yearly_assets():
    """CSV 파일에서 연말 자산가치 로드 (주식계좌 + 저축연금 + NH퇴직연금)."""
    import os
    csv_dir = os.path.join(DATA_DIR, 'raw_data', 'originals', '현재')
    asset = {}

    # 키움 수익률현황 스크린샷 기반 (연말 예탁자산)
    kiwoom_assets = {
        '2019': 457239,
        '2020': 28071390,
        '2021': 71135833,
        '2022': 44244243,
        '2023': 0,  # 전량 대체출고
    }
    for y, v in kiwoom_assets.items():
        asset.setdefault(y, {})['kiwoom'] = v

    # 2023.csv: col16=2022-12, col4=2023-12
    p = os.path.join(csv_dir, '2023.csv')
    if os.path.exists(p):
        with open(p, encoding='utf-8') as f:
            for row in csv.reader(f):
                if len(row) < 17: continue
                n = row[1].strip()
                if n == '키움증권':
                    # 2023.12 CSV "키움증권" = 실제로는 토스 (키움→토스 이관 완료 후)
                    asset.setdefault('2023', {})['toss'] = _parse_krw_cell(row[4])
                if n == '예슬 저축연금':
                    # 2022년 저축연금: 신한/미래에셋/카카오페이 2022년 이전 거래 없음 → 제외
                    asset.setdefault('2023', {})['pension'] = _parse_krw_cell(row[4])

    # 2024.csv: col4=2024-12
    p = os.path.join(csv_dir, '2024.csv')
    if os.path.exists(p):
        with open(p, encoding='utf-8') as f:
            for row in csv.reader(f):
                if len(row) < 17: continue
                n = row[1].strip()
                if n == '토스증권':
                    asset.setdefault('2024', {})['toss'] = _parse_krw_cell(row[4])
                if n == '예슬 저축연금':
                    asset.setdefault('2024', {})['pension'] = _parse_krw_cell(row[4])

    # 2025.csv: col4=2025-12
    p = os.path.join(csv_dir, '2025.csv')
    if os.path.exists(p):
        with open(p, encoding='utf-8') as f:
            for row in csv.reader(f):
                if len(row) < 5: continue
                n = row[1].strip()
                if n == '토스증권':
                    asset.setdefault('2025', {})['toss'] = _parse_krw_cell(row[4])
                if n == '예슬 저축연금':
                    asset.setdefault('2025', {})['pension'] = _parse_krw_cell(row[4])

    # 2026: 스크린샷 기반 (2026-03-23)
    asset['2026'] = {
        'toss': 148645505,       # 토스증권
        'pension': 16765064,     # 신한 9,208,573 + 카카오페이 6,026,656 + 미래에셋 1,529,835
        'nh_irp': 65531274,      # NH 퇴직연금신탁
        'snapshot_date': '2026-03-23',
    }

    # stock = kiwoom + toss (키움→토스 이관, 시기별 합산)
    for y in asset:
        asset[y].setdefault('kiwoom', 0)
        asset[y].setdefault('toss', 0)
        asset[y].setdefault('pension', 0)
        asset[y].setdefault('nh_irp', 0)
        asset[y]['stock'] = asset[y]['kiwoom'] + asset[y]['toss']
        asset[y]['total'] = asset[y]['stock'] + asset[y]['pension'] + asset[y]['nh_irp']

    return asset


def write_summary(trades, deps, path):
    yearly_assets = _load_yearly_assets()
    fmt, cfmt, cfmt_roi = _fmt_helpers()

    # 비바리퍼블리카 제외
    trades = [t for t in trades if t.get('stock') != '비바리퍼블리카']

    # 수동 입금 추가
    deps = list(deps) + [
        {'date': '2026-03-12', 'type': 'deposit', 'amount': 65071700,
         'broker': 'NH투자증권', 'description': '퇴직금입금'},
        {'date': '2020-11-04', 'type': 'deposit', 'amount': 1992000,
         'broker': '토스', 'description': '타증권사입고(주식현물)'},
    ]

    lines = ['# 투자 자산 요약\n']

    # ══════════════════════════════════════
    # 섹션 1: 연간 거래 요약 (전체)
    # ══════════════════════════════════════
    lines.append('## 1. 연간 거래 요약 (전체)\n')
    lines.append('> 예탁자산 = 주식계좌(키움/토스) + 저축연금 + NH퇴직연금')
    lines.append('> 손익 = 자산변동 - 순입금 (실현+미실현)\n')

    # 연도별 입출금 집계
    dep_by_year = defaultdict(lambda: {'dep': 0, 'wd': 0})
    for d in deps:
        dep_by_year[d['date'][:4]]['dep' if d['type'] == 'deposit' else 'wd'] += d['amount']

    all_years = sorted(set(list(dep_by_year.keys()) + list(yearly_assets.keys())))

    lines.append('| 연도 | 예탁자산 | 입금 | 출금 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|---------|------|------|------|-------|----------|---------|')
    cum_dep, cum_pnl, prev_asset = 0, 0, 0
    for y in all_years:
        v = dep_by_year[y]
        net = v['dep'] - v['wd']; cum_dep += net
        a = yearly_assets.get(y, {}).get('total')
        if a and prev_asset > 0:
            pnl = a - prev_asset - net; cum_pnl += pnl
            roi = (pnl / prev_asset * 100) if prev_asset else 0
            pnl_s = cfmt(int(pnl)); roi_s = cfmt_roi(roi)
        elif a and prev_asset == 0:
            cum_pnl = a - cum_dep
            pnl_s = '-'; roi_s = '-'
        else:
            pnl_s = '-'; roi_s = '-'
        if a: prev_asset = a
        asset_s = fmt(a) if a else '-'
        dep_s = fmt(v['dep']) if v['dep'] else '-'
        wd_s = fmt(v['wd']) if v['wd'] else '-'
        lines.append(f'| {y} | {asset_s} | {dep_s} | {wd_s} | {pnl_s} | {roi_s} | {fmt(cum_dep)} | {cfmt(int(cum_pnl))} |')
    lines.append('')

    # 연말 자산 내역
    lines.append('### 연말 자산 내역\n')
    lines.append('| 연도 | 키움 | 토스 | 주식소계 | 저축연금 | NH퇴직연금 | 합계 |')
    lines.append('|------|------|------|---------|---------|-----------|------|')
    for y in sorted(yearly_assets.keys()):
        a = yearly_assets[y]
        kw = fmt(a['kiwoom']) if a['kiwoom'] else '-'
        ts = fmt(a['toss']) if a['toss'] else '-'
        pn = fmt(a['pension']) if a['pension'] else '-'
        nh = fmt(a['nh_irp']) if a['nh_irp'] else '-'
        lines.append(f'| {y} | {kw} | {ts} | {fmt(a["stock"])} | {pn} | {nh} | {fmt(a["total"])} |')
    lines.append('')

    # ══════════════════════════════════════
    # 섹션 2: 주식 거래요약 (키움 + 토스)
    # ══════════════════════════════════════
    lines.append('---\n')
    lines.append('## 2. 주식 거래요약 (키움 + 토스)\n')

    # 주식 수익률 해석 먼저
    lines.append('### 주식 수익률 해석\n')
    lines.append('| 연도 | 계좌 | 수익률 | 주요 종목/이벤트 |')
    lines.append('|------|------|-------|--------------|')
    stock_interp = [
        ('2019', '키움', -0.06, '12월 시작. AT&T·스타벅스 소액 진입'),
        ('2020', '키움', 34.19, '코로나 저점 매수. 테슬라(727만)·알파벳·애플·엔비디아 진입'),
        ('2021', '키움', 40.20, '테슬라 집중(910만). 아마존·SPY 확대'),
        ('2022', '키움', -39.79, '금리인상 하락장. 전 종목 하락'),
        ('2023', '키움→토스', 0, '키움→토스 대체출고 7165만원. 엔비디아·DIA·마이크로소프트 추가'),
        ('2024', '토스', 88.77, '엔비디아 급등 수혜. NVDY·JEPI 배당전략. 애플 매도(1963만)'),
        ('2025', '토스', 17.90, 'NVDY·JEPI 정리. 테슬라·알파벳 추가매수. 카카오페이/케이뱅크 출금'),
        ('2026', '토스', -10.40, '마이크로소프트 매도. 알파벳 집중매수(404만). 연초 조정'),
    ]
    for y, acct, roi, comment in stock_interp:
        lines.append(f'| {y} | {acct} | {cfmt_roi(roi)} | {comment} |')
    lines.append('')

    # 키움 요약
    kiwoom_data = [
        (2019, 147010, 0, 457239, -85, -0.06),
        (2020, 19392987, 0, 28071390, 6786798, 34.19),
        (2021, 23241920, 120565, 71135833, 20581331, 40.20),
        (2022, 3928279, 0, 44244243, -29868848, -39.79),
        (2023, 0, 66658, 0, 27420991, 0.00),
    ]
    lines.append('### 키움증권 (2019~2023, 폐쇄)\n')
    lines.append('| 연도 | 입금 | 출금 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|------|------|------------|------|-------|----------|---------|')
    ck, cpk = 0, 0
    for y, dep, wd, asset, pnl, roi in kiwoom_data:
        ck += dep - wd; cpk += pnl
        lines.append(f'| {y} | {fmt(dep) if dep else "-"} | {fmt(wd) if wd else "-"} | {fmt(asset) if asset else "-"} | {cfmt(pnl)} | {cfmt_roi(roi)} | {fmt(ck)} | {cfmt(cpk)} |')
    lines.append(f'\n> 상세: [키움증권.md](account/키움증권.md)\n')

    # 토스 요약
    toss_deps_list = [d for d in deps if d['broker'] == '토스']
    toss_by_year = defaultdict(lambda: {'dep': 0, 'wd': 0})
    for d in toss_deps_list:
        toss_by_year[d['date'][:4]]['dep' if d['type'] == 'deposit' else 'wd'] += d['amount']
    toss_years = sorted(set(list(toss_by_year.keys()) +
                            [y for y in yearly_assets if yearly_assets[y].get('toss', 0) > 0]))

    KIWOOM_TRANSFER = 71649350
    lines.append('### 토스증권 (2021~현재, 주력)\n')
    lines.append('> 2023.07 키움→토스 주식 입고 7165만원 포함\n')
    lines.append('| 연도 | 현금입금 | 현금출금 | 주식입고 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|---------|---------|---------|------------|------|-------|----------|---------|')
    ct, cpt, prev_t = 0, 0, 0
    for y in toss_years:
        v = toss_by_year[y]
        stock_in = KIWOOM_TRANSFER if y == '2023' else 0
        net_cash = v['dep'] - v['wd']
        ct += net_cash + stock_in
        a = yearly_assets.get(y, {}).get('toss', 0)
        if a and prev_t > 0:
            pnl = a - prev_t - net_cash - stock_in; cpt += pnl
            roi = (pnl / prev_t * 100) if prev_t else 0
            pnl_s = cfmt(int(pnl)); roi_s = cfmt_roi(roi)
        else:
            pnl_s = '-'; roi_s = '-'
        if a: prev_t = a
        stock_s = fmt(stock_in) if stock_in else '-'
        lines.append(f'| {y} | {fmt(v["dep"]) if v["dep"] else "-"} | {fmt(v["wd"]) if v["wd"] else "-"} | {stock_s} | {fmt(a) if a else "-"} | {pnl_s} | {roi_s} | {fmt(ct)} | {cfmt(int(cpt)) if cpt else "-"} |')
    lines.append(f'\n> 상세: [토스증권.md](account/토스증권.md)\n')

    # ══════════════════════════════════════
    # 섹션 3: 저축연금 + 퇴직연금
    # ══════════════════════════════════════
    lines.append('---\n')
    lines.append('## 3. 연금 (저축연금 + 퇴직연금)\n')

    # 연금 전체 합산 (저축연금 + NH 퇴직연금)
    nh_dep_s = 65071700; nh_asset_s = 65531274; nh_pnl_s = nh_asset_s - nh_dep_s
    pension_yearly = {
        '2023': (5010000, 10000, 5000000, 0),
        '2024': (10000, 0, 8008713, 1998713),
        '2025': (6000000, 0, 15461213, 1452500),
        '2026': (200000 + nh_dep_s, 0, 16562842 + nh_asset_s, 901629 + nh_pnl_s),
    }

    lines.append('### 연금 연간 요약\n')
    lines.append('| 연도 | 입금 | 출금 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|------|------|------------|------|-------|----------|---------|')
    cp_dep, cp_pnl, prev_pen = 0, 0, 0
    for y in sorted(pension_yearly.keys()):
        dep, wd, asset, pnl = pension_yearly[y]
        cp_dep += dep - wd; cp_pnl += pnl
        roi = (pnl / prev_pen * 100) if prev_pen > 0 else 0.0
        lines.append(f'| {y} | {fmt(dep) if dep else "-"} | {fmt(wd) if wd else "-"} | {fmt(asset)} | {cfmt(pnl)} | {cfmt_roi(roi) if prev_pen > 0 else "-"} | {fmt(cp_dep)} | {cfmt(cp_pnl)} |')
        prev_pen = asset
    lines.append('')

    # 연금 수익률 해석 (요약 바로 밑)
    lines.append('### 연금 수익률 해석\n')
    lines.append('| 연도 | 계좌 | 수익률 | 주요 종목/이벤트 |')
    lines.append('|------|------|-------|--------------|')
    pension_interp = [
        ('2023', '신한', 0, '신한 개설. 500만원 입금. 미국나스닥100·S&P500·인도니프티50 ETF 시작'),
        ('2024', '신한+미래', 39.97, '신한 +35%(미국·인도 ETF 강세). 미래에셋 +24%(인도니프티50 수혜)'),
        ('2025', '3개 계좌', 18.14, '카카오페이 개설(470만). 신한 +19%(리밸런싱). 미래 +2.7%'),
        ('2026', '4개 계좌', 5.83, '카카오페이 +20%(배당ETF 강세). 신한 -1.9%. NH 퇴직금 6507만 입금 (정기예금)'),
    ]
    for y, acct, roi, comment in pension_interp:
        lines.append(f'| {y} | {acct} | {cfmt_roi(roi)} | {comment} |')
    lines.append('')

    # 계좌별 현황
    nh_dep = 65071700; nh_asset = 65531274; nh_pnl = nh_asset - nh_dep

    lines.append('### 계좌별 현황\n')
    lines.append('| 계좌 | 2026.03 기준 | 누적수익률 |')
    lines.append('|------|------------|----------|')
    lines.append(f'| 신한 연금저축 | ₩9,070,213 | {cfmt_roi(57.41)} |')
    lines.append(f'| 카카오페이 연금저축 | ₩6,028,474 | {cfmt_roi(23.03)} |')
    lines.append(f'| 미래에셋 연금저축(신) | ₩1,464,155 | {cfmt_roi(31.93)} |')
    lines.append(f'| NH 퇴직연금 (IRP) | {fmt(nh_asset)} | {cfmt_roi(nh_pnl / nh_dep * 100)} |')
    lines.append(f'| **합계** | **{fmt(16562842 + nh_asset)}** | |')
    lines.append(f'\n> 상세: [신한](account/신한_연금저축.md) · [카카오페이](account/카카오페이_연금저축.md) · [미래에셋](account/미래에셋_연금저축.md) · [NH](account/NH_퇴직연금.md)\n')

    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote summary to {path}")


def _fmt_helpers():
    """공통 포매팅 헬퍼 반환."""
    fmt = lambda n: f"-₩{abs(n):,}" if n < 0 else f"₩{n:,}"
    cfmt = lambda n: f'<span style="color:#e74c3c">{fmt(n)}</span>' if n > 0 else (f'<span style="color:#3498db">{fmt(n)}</span>' if n < 0 else fmt(n))
    cfmt_roi = lambda r: f'<span style="color:#e74c3c">{r:.2f}%</span>' if r > 0 else (f'<span style="color:#3498db">{r:.2f}%</span>' if r < 0 else f'{r:.2f}%')
    return fmt, cfmt, cfmt_roi


def _write_deposit_detail(lines, broker_deps, fmt):
    """입출금 상세내역을 연도별로 작성."""
    lines.append('## 입출금 상세내역\n')
    sorted_deps = sorted(broker_deps, key=lambda x: x['date'])
    if not sorted_deps:
        lines.append('> 입출금 내역 없음\n')
        return
    by_year = defaultdict(list)
    for d in sorted_deps:
        by_year[d['date'][:4]].append(d)
    for y in sorted(by_year.keys()):
        yd = by_year[y]
        y_dep = sum(d['amount'] for d in yd if d['type'] == 'deposit')
        y_wd = sum(d['amount'] for d in yd if d['type'] == 'withdrawal')
        parts = []
        if y_dep: parts.append(f'입금 {fmt(y_dep)}')
        if y_wd: parts.append(f'출금 {fmt(y_wd)}')
        lines.append(f'### {y}년 ({" / ".join(parts) if parts else "거래 없음"})\n')
        lines.append('| 날짜 | 구분 | 금액 | 설명 |')
        lines.append('|------|------|------|------|')
        for d in yd:
            typ = '입금' if d['type'] == 'deposit' else '출금'
            lines.append(f'| {d["date"]} | {typ} | {fmt(d["amount"])} | {d.get("description","")} |')
        lines.append('')
    lines.append(f'> 총 {len(sorted_deps)}건\n')


def write_account_md(deps, output_dir):
    """증권사별 계좌 상세 MD 파일 생성."""
    import os
    account_dir = os.path.join(output_dir, 'account')
    os.makedirs(account_dir, exist_ok=True)
    fmt, cfmt, cfmt_roi = _fmt_helpers()
    yearly_assets = _load_yearly_assets()

    # ══════════════════════════════════════
    # 키움증권
    # ══════════════════════════════════════
    kiwoom_data = [
        # (연도, 입금, 출금, 예탁자산, 평가금, 손익, 수익률%)
        (2019, 147010, 0, 457239, 147041, -85, -0.06),
        (2020, 19392987, 0, 28071390, 25053245, 6786798, 34.19),
        (2021, 23241920, 120565, 71135833, 70060632, 20581331, 40.20),
        (2022, 3928279, 0, 44244243, 44093384, -29868848, -39.79),
        (2023, 0, 66658, 0, 0, 27420991, 0.00),
    ]
    kiwoom_comments = {
        2019: '12월 시작, 워밍업',
        2020: '코로나 저점 매수 → 반등. 테슬라(727만)·알파벳·애플·엔비디아 진입. 1940만원 투입',
        2021: '테슬라 집중 매수(910만). 아마존·SPY·DIA 확대. 2320만원 투입',
        2022: '금리인상 하락장. 전 종목 하락. 2년 수익 전부 증발',
        2023: '키움→토스 대체출고. 출고 시 실현손익 반영',
    }

    lines = ['# 키움증권\n']
    lines.append('> 계좌: 5289-3356 [위탁종합] 이예슬')
    lines.append('> 기간: 2019.12 ~ 2023.07 (키움→토스 대체출고 후 폐쇄)')
    lines.append('> 데이터: 키움 수익률현황 스크린샷 + PDF 거래내역서\n')

    lines.append('## 연간 요약\n')
    lines.append('| 연도 | 입금 | 출금 | 연말 예탁자산 | 연말 주식 평가금 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|------|------|------------|-------------|------|-------|----------|---------|')
    cum_dep, cum_pnl = 0, 0
    rows = []
    for y, dep, wd, asset, ev, pnl, roi in kiwoom_data:
        cum_dep += dep - wd; cum_pnl += pnl
        cum_roi = (cum_pnl / cum_dep * 100) if cum_dep > 0 else 0
        lines.append(f'| {y} | {fmt(dep) if dep else "-"} | {fmt(wd) if wd else "-"} | {fmt(asset) if asset else "-"} | {fmt(ev) if ev else "-"} | {cfmt(pnl)} | {cfmt_roi(roi)} | {fmt(cum_dep)} | {cfmt(cum_pnl)} |')
        rows.append((y, roi, cum_roi))
    lines.append('')

    lines.append('## 수익률 해석\n')
    lines.append('| 연도 | 연간수익률 | 누적수익률 | 해석 |')
    lines.append('|------|----------|----------|------|')
    for y, roi, cum_roi in rows:
        lines.append(f'| {y} | {cfmt_roi(roi)} | {cfmt_roi(cum_roi)} | {kiwoom_comments.get(y, "")} |')

    total_dep = sum(d for _,d,_,_,_,_,_ in kiwoom_data)
    total_wd = sum(w for _,_,w,_,_,_,_ in kiwoom_data)
    ann_roi = cum_roi / len(kiwoom_data)
    lines.append('')
    lines.append(f'> **총 투입**: {fmt(int(total_dep))} | **총 출금**: {fmt(int(total_wd))} | **순투입**: {fmt(int(total_dep - total_wd))}')
    lines.append(f'> **최종 손익**: {cfmt(int(cum_pnl))} | **누적수익률**: {cfmt_roi(cum_roi)} | **연평균**: ~{cfmt_roi(ann_roi)}')
    lines.append('')

    _write_deposit_detail(lines, [d for d in deps if d['broker'] == '키움'], fmt)
    with open(os.path.join(account_dir, '키움증권.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote account/키움증권.md")

    # ══════════════════════════════════════
    # 토스증권
    # ══════════════════════════════════════
    toss_deps = [d for d in deps if d['broker'] == '토스']
    toss_by_year = defaultdict(lambda: {'dep': 0, 'wd': 0})
    for d in toss_deps:
        toss_by_year[d['date'][:4]]['dep' if d['type'] == 'deposit' else 'wd'] += d['amount']

    toss_years = sorted(set(list(toss_by_year.keys()) +
                            [y for y in yearly_assets if yearly_assets[y].get('toss', 0) > 0]))

    # 키움→토스 주식 입고 금액 (키움 스크린샷 2023 출고: ₩71,649,350)
    KIWOOM_STOCK_TRANSFER = 71649350  # 2023.07 키움→토스 대체출고

    lines = ['# 토스증권\n']
    lines.append('> 기간: 2021.01 ~ 현재 (주력 주식계좌)')
    lines.append('> 2023.07 키움→토스 대체출고 (7165만원 상당) 이후 전체 주식 운용\n')

    lines.append('## 연간 요약\n')
    lines.append('| 연도 | 현금입금 | 현금출금 | 주식입고 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|---------|---------|---------|------------|------|-------|----------|---------|')
    cum_dep_t, cum_pnl_t, prev_asset_t = 0, 0, 0
    toss_rows = []
    for y in toss_years:
        v = toss_by_year[y]
        stock_in = KIWOOM_STOCK_TRANSFER if y == '2023' else 0
        net_cash = v['dep'] - v['wd']
        cum_dep_t += net_cash + stock_in
        a = yearly_assets.get(y, {}).get('toss', 0)
        if a and prev_asset_t > 0:
            pnl = a - prev_asset_t - net_cash - stock_in
            roi = (pnl / prev_asset_t * 100) if prev_asset_t else 0
            cum_pnl_t += pnl
        elif a and prev_asset_t == 0:
            pnl = None; roi = None
        else:
            pnl = None; roi = None
        if a: prev_asset_t = a

        dep_s = fmt(v['dep']) if v['dep'] else '-'
        wd_s = fmt(v['wd']) if v['wd'] else '-'
        stock_s = fmt(stock_in) if stock_in else '-'
        asset_s = fmt(a) if a else '-'
        pnl_s = cfmt(int(pnl)) if pnl is not None else '-'
        roi_s = cfmt_roi(roi) if roi is not None else '-'
        cum_roi_t = (cum_pnl_t / cum_dep_t * 100) if cum_dep_t > 0 else 0
        cum_pnl_s = cfmt(int(cum_pnl_t)) if cum_pnl_t else '-'
        lines.append(f'| {y} | {dep_s} | {wd_s} | {stock_s} | {asset_s} | {pnl_s} | {roi_s} | {fmt(cum_dep_t)} | {cum_pnl_s} |')
        toss_rows.append((y, pnl, roi, cum_pnl_t, cum_dep_t, cum_roi_t))
    lines.append('')

    toss_comments = {
        '2023': '키움→토스 주식 대체입고 7165만원. 엔비디아(312만)·DIA·마이크로소프트 추가 매수',
        '2024': '엔비디아 급등 수혜(631만 추가매수). NVDY·JEPI 배당전략 활발. 애플 일부 매도(1963만)',
        '2025': 'NVDY(2107만 매도)·JEPI(1131만 매도) 정리. 테슬라·알파벳 추가매수. 카카오페이/케이뱅크 출금',
        '2026': '마이크로소프트 매도(422만). 알파벳 집중 매수(404만). 연초 시장 조정 -10%',
    }
    has_roi = any(r is not None for _, _, r, _, _, _ in toss_rows)
    if has_roi:
        lines.append('## 수익률 해석\n')
        lines.append('| 연도 | 연간수익률 | 누적수익률 | 해석 |')
        lines.append('|------|----------|----------|------|')
        for y, pnl, roi, cum_pnl_v, cum_dep_v, cum_roi_v in toss_rows:
            if roi is not None:
                lines.append(f'| {y} | {cfmt_roi(roi)} | {cfmt_roi(cum_roi_v)} | {toss_comments.get(y, "")} |')

        # 요약 통계
        total_cash = sum(toss_by_year[y]['dep'] for y in toss_years)
        total_wd = sum(toss_by_year[y]['wd'] for y in toss_years)
        ann_t = cum_roi_t / len([r for r in toss_rows if r[1] is not None]) if toss_rows else 0
        lines.append('')
        lines.append(f'> **현금 투입**: {fmt(int(total_cash))} | **현금 출금**: {fmt(int(total_wd))} | **주식 입고**: {fmt(KIWOOM_STOCK_TRANSFER)}')
        lines.append(f'> **순투입 합계**: {fmt(int(cum_dep_t))} | **최종 손익**: {cfmt(int(cum_pnl_t))} | **누적수익률**: {cfmt_roi(cum_roi_t)}')
        lines.append('')

    _write_deposit_detail(lines, toss_deps, fmt)
    with open(os.path.join(account_dir, '토스증권.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote account/토스증권.md")

    # ══════════════════════════════════════
    # 연금저축 계좌들 (신한/카카오페이/미래에셋) — 동일 패턴
    # ══════════════════════════════════════
    # ══════════════════════════════════════
    # 신한 연금저축
    # ══════════════════════════════════════
    shinhan_data = [
        # (연도, 입금, 출금, 연말자산, 손익, 수익률%)
        (2023, 5010000, 10000, 5000000, 0, 0.00),
        (2024, 0, 0, 6754535, 1754535, 35.09),
        (2025, 1200000, 0, 9247961, 1293426, 19.15),
        (2026, 0, 0, 9070213, -177748, -1.92),  # 3/23 기준
    ]
    shinhan_comments = {
        2023: '12월 개설. 500만원 입금. TIGER 미국나스닥100·ACE 미국S&P500·TIGER 인도니프티50 매수',
        2024: '추가 입금 없이 시장 수익만 +35%. 미국·인도 ETF 강세',
        2025: '120만원 추가 입금. 11월 포트 리밸런싱 (인도·배당성장 매도 → 나스닥·S&P500 집중)',
        2026: '연초 조정 -1.9%. 3/23 기준',
    }
    lines = ['# 신한 연금저축\n']
    lines.append('> 기간: 2023.12 ~ 현재')
    lines.append('> 주요 보유: TIGER 미국나스닥100, ACE 미국S&P500, TIGER 배당성장\n')

    lines.append('## 연간 요약\n')
    lines.append('| 연도 | 입금 | 출금 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|------|------|------------|------|-------|----------|---------|')
    cs, cps = 0, 0
    sh_rows = []
    for y, dep, wd, asset, pnl, roi in shinhan_data:
        cs += dep - wd; cps += pnl
        cr = (cps / cs * 100) if cs > 0 else 0
        lines.append(f'| {y} | {fmt(dep) if dep else "-"} | {fmt(wd) if wd else "-"} | {fmt(asset)} | {cfmt(pnl)} | {cfmt_roi(roi)} | {fmt(cs)} | {cfmt(cps)} |')
        sh_rows.append((y, roi, cr))
    lines.append('')

    lines.append('## 수익률 해석\n')
    lines.append('| 연도 | 연간수익률 | 누적수익률 | 해석 |')
    lines.append('|------|----------|----------|------|')
    for y, roi, cr in sh_rows:
        lines.append(f'| {y} | {cfmt_roi(roi)} | {cfmt_roi(cr)} | {shinhan_comments.get(y, "")} |')
    total_dep_sh = sum(d for _,d,_,_,_,_ in shinhan_data)
    total_wd_sh = sum(w for _,_,w,_,_,_ in shinhan_data)
    ann_sh = cr / len([r for r in sh_rows if r[1] != 0]) if sh_rows else 0
    lines.append('')
    lines.append(f'> **순투입**: {fmt(int(total_dep_sh - total_wd_sh))} | **최종 손익**: {cfmt(int(cps))} | **누적수익률**: {cfmt_roi(cr)}')
    lines.append('')

    # 현재 보유종목
    lines.append('## 현재 보유종목 (2026.03.23 기준)\n')
    lines.append(f'> 계좌: 신한투자증권 27071883501')
    lines.append(f'> 총 평가: ₩9,092,023 (현금 ₩25,273)\n')
    lines.append('| 종목 | 수량 | 평가금액 | 수익 | 수익률 |')
    lines.append('|------|------|---------|------|-------|')
    lines.append(f'| TIGER 미국나스닥100 | 30주 | ₩4,782,000 | {cfmt(1000975)} | {cfmt_roi(26.5)} |')
    lines.append(f'| ACE 미국S&P500 | 174주 | ₩4,284,750 | {cfmt(961730)} | {cfmt_roi(28.9)} |')
    lines.append('')

    _write_deposit_detail(lines, [d for d in deps if d['broker'] == '신한'], fmt)
    with open(os.path.join(account_dir, '신한_연금저축.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote account/신한_연금저축.md")

    # ══════════════════════════════════════
    # 카카오페이 연금저축
    # ══════════════════════════════════════
    kakao_data = [
        (2025, 4700000, 0, 4824549, 124549, 2.65),
        (2026, 200000, 0, 6028474, 1003925, 19.98),  # 3/22 기준
    ]
    kakao_comments = {
        2025: '11월 개설. 토스에서 470만원 이체. SOL 미국배당(179만)·TIGER 배당성장(140만)·ACE 미국배당(117만) 일괄 매수',
        2026: '20만원 추가 납입. 배당 ETF 강세로 +19.98%. 월 배당금 유입 시작',
    }
    lines = ['# 카카오페이 연금저축\n']
    lines.append('> 계좌: 020-09-790094')
    lines.append('> 기간: 2025.11 ~ 현재')
    lines.append('> 주요 보유: SOL 미국배당다우존스, TIGER 배당성장, ACE 미국배당퀄리티\n')

    lines.append('## 연간 요약\n')
    lines.append('| 연도 | 입금 | 출금 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|------|------|------------|------|-------|----------|---------|')
    ck2, cpk2 = 0, 0
    kk_rows = []
    for y, dep, wd, asset, pnl, roi in kakao_data:
        ck2 += dep - wd; cpk2 += pnl
        cr2 = (cpk2 / ck2 * 100) if ck2 > 0 else 0
        lines.append(f'| {y} | {fmt(dep) if dep else "-"} | {fmt(wd) if wd else "-"} | {fmt(asset)} | {cfmt(pnl)} | {cfmt_roi(roi)} | {fmt(ck2)} | {cfmt(cpk2)} |')
        kk_rows.append((y, roi, cr2))
    lines.append('')

    lines.append('## 수익률 해석\n')
    lines.append('| 연도 | 연간수익률 | 누적수익률 | 해석 |')
    lines.append('|------|----------|----------|------|')
    for y, roi, cr2 in kk_rows:
        lines.append(f'| {y} | {cfmt_roi(roi)} | {cfmt_roi(cr2)} | {kakao_comments.get(y, "")} |')
    lines.append('')
    lines.append(f'> **순투입**: {fmt(int(ck2))} | **최종 손익**: {cfmt(int(cpk2))} | **누적수익률**: {cfmt_roi(cr2)}')
    lines.append('')

    # 현재 보유종목
    lines.append('## 현재 보유종목 (2026.03.23 기준)\n')
    lines.append(f'> 계좌: 카카오페이증권 02009790094')
    lines.append(f'> 총 평가: ₩5,845,647 (현금 ₩24,297)\n')
    lines.append('| 종목 | 수량 | 평가금액 | 수익 | 수익률 |')
    lines.append('|------|------|---------|------|-------|')
    lines.append(f'| SOL 미국배당다우존스 | 156주 | ₩2,075,580 | {cfmt(257110)} | {cfmt_roi(14.1)} |')
    lines.append(f'| TIGER 배당성장 | 57주 | ₩1,993,290 | {cfmt(492585)} | {cfmt_roi(32.8)} |')
    lines.append(f'| ACE 미국배당퀄리티 | 101주 | ₩1,217,050 | {cfmt(30680)} | {cfmt_roi(2.6)} |')
    lines.append(f'| TIGER 배당커버드콜액티브 | 27주 | ₩460,215 | {cfmt(114015)} | {cfmt_roi(32.9)} |')
    lines.append(f'| KODEX 골드선물(H) | 2주 | ₩48,400 | {cfmt(-6830)} | {cfmt_roi(-12.4)} |')
    lines.append(f'| KODEX AI반도체 | 1주 | ₩26,815 | {cfmt(-365)} | {cfmt_roi(-1.3)} |')
    lines.append('')

    _write_deposit_detail(lines, [d for d in deps if d['broker'] == '카카오페이'], fmt)
    with open(os.path.join(account_dir, '카카오페이_연금저축.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote account/카카오페이_연금저축.md")

    # ══════════════════════════════════════
    # 미래에셋 연금저축(신)
    # ══════════════════════════════════════
    mirae_data = [
        (2024, 10000, 0, 1254178, 244178, 24.21),
        (2025, 100000, 0, 1388703, 34525, 2.65),
        (2026, 0, 0, 1464155, 75452, 5.43),  # 현재
    ]
    mirae_comments = {
        2024: '100만원으로 시작 (2023말 입금). TIGER 인도니프티50·KODEX 미국S&P500TR 매수. 인도 강세 수혜',
        2025: '10만원 추가 납입. 시장 보합으로 소폭 수익',
        2026: '추가 납입 없이 +5.43%. 현재 기준',
    }
    lines = ['# 미래에셋 연금저축(신)\n']
    lines.append('> 기간: 2023말 ~ 현재')
    lines.append('> 주요 보유: TIGER 인도니프티50, KODEX 미국S&P500TR\n')

    lines.append('## 연간 요약\n')
    lines.append('| 연도 | 입금 | 출금 | 연말 예탁자산 | 손익 | 수익률 | 누적순입금 | 누적손익 |')
    lines.append('|------|------|------|------------|------|-------|----------|---------|')
    cm2, cpm = 0, 0
    mr_rows = []
    for y, dep, wd, asset, pnl, roi in mirae_data:
        cm2 += dep - wd; cpm += pnl
        cr3 = (cpm / (cm2 + 1000000) * 100) if (cm2 + 1000000) > 0 else 0  # 초기 100만 포함
        lines.append(f'| {y} | {fmt(dep) if dep else "-"} | {fmt(wd) if wd else "-"} | {fmt(asset)} | {cfmt(pnl)} | {cfmt_roi(roi)} | {fmt(cm2 + 1000000)} | {cfmt(cpm)} |')
        mr_rows.append((y, roi, cr3))
    lines.append('')

    lines.append('## 수익률 해석\n')
    lines.append('| 연도 | 연간수익률 | 누적수익률 | 해석 |')
    lines.append('|------|----------|----------|------|')
    for y, roi, cr3 in mr_rows:
        lines.append(f'| {y} | {cfmt_roi(roi)} | {cfmt_roi(cr3)} | {mirae_comments.get(y, "")} |')
    lines.append('')
    lines.append(f'> **순투입**: {fmt(int(cm2 + 1000000))} | **최종 손익**: {cfmt(int(cpm))} | **누적수익률**: {cfmt_roi(cr3)}')
    lines.append('')

    # 현재 보유종목
    lines.append('## 현재 보유종목 (2026.03.23 기준)\n')
    lines.append(f'> 계좌: 미래에셋증권 010753582810')
    lines.append(f'> 총 평가: ₩1,466,015 (현금 ₩232,515)\n')
    lines.append('| 종목 | 수량 | 평가금액 | 수익 | 수익률 |')
    lines.append('|------|------|---------|------|-------|')
    lines.append(f'| KODEX 200 타겟위클리커버드콜 | 40주 | ₩674,600 | {cfmt(165755)} | {cfmt_roi(32.6)} |')
    lines.append(f'| TIGER 인도니프티50 | 46주 | ₩558,900 | {cfmt(-107870)} | {cfmt_roi(-16.2)} |')
    lines.append('')

    _write_deposit_detail(lines, [d for d in deps if d['broker'] == '미래에셋'], fmt)
    with open(os.path.join(account_dir, '미래에셋_연금저축.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote account/미래에셋_연금저축.md")

    # ══════════════════════════════════════
    # NH 퇴직연금
    # ══════════════════════════════════════
    lines = ['# NH투자증권 퇴직연금 (IRP)\n']
    lines.append('> 계좌: 21102511034')
    lines.append('> 2026.03.12 퇴직금 입금으로 개설\n')

    lines.append('## 연간 요약\n')
    lines.append('| 연도 | 입금 | 연말 예탁자산 | 손익 | 수익률 |')
    lines.append('|------|------|------------|------|-------|')
    nh_asset = 65531274; nh_dep = 65071700; nh_pnl = nh_asset - nh_dep
    nh_roi = (nh_pnl / nh_dep * 100)
    lines.append(f'| 2026 | {fmt(nh_dep)} | {fmt(nh_asset)} | {cfmt(nh_pnl)} | {cfmt_roi(nh_roi)} |')
    lines.append('')

    lines.append('## 수익률 해석\n')
    lines.append('| 연도 | 연간수익률 | 해석 |')
    lines.append('|------|----------|------|')
    lines.append(f'| 2026 | {cfmt_roi(nh_roi)} | 퇴직금 6507만원 입금. 정기예금 중심 안전 운용 (1Y 예금 2건) |')
    lines.append('')

    lines.append('## 자산 구성 (2026.03.23 기준)\n')
    lines.append('| 상품 | 금액 |')
    lines.append('|------|------|')
    lines.append('| 자동운용상품(고유계정대) | ₩50,105,923 |')
    lines.append('| 한국투자저축은행 정기예금_1Y [개인IRP] | ₩10,000,863 |')
    lines.append('| 키움저축은행 정기예금_1Y [개인IRP] | ₩5,000,431 |')
    lines.append(f'| **합계** | **{fmt(nh_asset)}** |')
    lines.append('')

    lines.append('## 입출금 상세내역\n')
    lines.append('### 2026년 (입금 ₩65,071,700)\n')
    lines.append('| 날짜 | 구분 | 금액 | 설명 |')
    lines.append('|------|------|------|------|')
    lines.append(f'| 2026-03-12 | 입금 | {fmt(nh_dep)} | 퇴직금입금 |')
    lines.append('\n> 총 1건\n')

    with open(os.path.join(account_dir, 'NH_퇴직연금.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Wrote account/NH_퇴직연금.md")


def filter_deposits(deps):
    filtered = []
    for d in deps:
        desc = d.get('description', '')
        # Exclude patterns
        if any(x in desc for x in ['타사대체','대체출고','대체입고','토스증권','키움증권',
                                    '전환입금','전환출금','배당금','분배금','예탁금이용료',
                                    '이자입금','매수출금','매도입금','환전','캐시백','대체TOSS']):
            continue
        # Include patterns
        if any(desc.startswith(t) for t in ['이체입금','이체출금','전자이체입금','전자이체출금',
                                             '이체오픈뱅킹','오픈뱅킹입금','연금저축정기납입',
                                             '대체입금(연금저축정기납입)','전자망이체입금']):
            filtered.append(d)
    return filtered


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("TRADE EXTRACTION FROM BROKER PDFs")
    print("=" * 60)

    all_trades, all_deposits, all_exrates = [], [], {}

    # KIWOOM
    for f in ['decrypted_키움_2019-2020.pdf','decrypted_키움_2021-2022.pdf','decrypted_키움_2023-2024.pdf']:
        fp = os.path.join(PARSED_DIR, f)
        if os.path.exists(fp):
            print(f"\nParsing {f}...")
            t, d, r = parse_kiwoom(fp)
            all_trades.extend(t); all_deposits.extend(d); all_exrates.update(r)
            print(f"  → {len(t)} trades, {len(d)} deposits, {len(r)} rates")

    print(f"\nComputing KRW for 키움 USD trades ({len(all_exrates)} rates)...")
    compute_kiwoom_krw(all_trades, all_exrates)

    # TOSS
    for f in ['decrypted_토스_2021.pdf','decrypted_토스_2022.pdf','decrypted_토스_2023.pdf',
              'decrypted_토스_2024.pdf','decrypted_토스_2025.pdf','decrypted_토스_2026.pdf']:
        fp = os.path.join(PARSED_DIR, f)
        if os.path.exists(fp):
            print(f"\nParsing {f}...")
            t, d = parse_toss(fp)
            all_trades.extend(t); all_deposits.extend(d)
            print(f"  → {len(t)} trades, {len(d)} deposits")

    # SHINHAN
    for f in ['decrypted_신한_2023.pdf','decrypted_신한_2024-2025.pdf','decrypted_신한_2025-2026.pdf']:
        fp = os.path.join(PARSED_DIR, f)
        if os.path.exists(fp):
            print(f"\nParsing {f}...")
            t, d = parse_shinhan(fp)
            all_trades.extend(t); all_deposits.extend(d)
            print(f"  → {len(t)} trades, {len(d)} deposits")

    # MIRAE
    for f in ['decrypted_미래에셋_1.pdf','decrypted_미래에셋_2.pdf']:
        fp = os.path.join(PARSED_DIR, f)
        if os.path.exists(fp):
            print(f"\nParsing {f}...")
            t, d = parse_mirae(fp)
            all_trades.extend(t); all_deposits.extend(d)
            print(f"  → {len(t)} trades, {len(d)} deposits")

    # KAKAOPAY
    fp = os.path.join(PARSED_DIR, 'decrypted_카카오페이_.pdf')
    if os.path.exists(fp):
        print(f"\nParsing 카카오페이...")
        t, d = parse_kakaopay(fp)
        all_trades.extend(t); all_deposits.extend(d)
        print(f"  → {len(t)} trades, {len(d)} deposits")

    # Deduplicate 미래에셋 only (_1 and _2 PDFs overlap)
    mirae_seen = set()
    deduped = []
    for t in all_trades:
        if t['broker'] == '미래에셋':
            key = (t['date'], t['type'], t['stock'], t['quantity'], t['amount'])
            if key in mirae_seen:
                continue
            mirae_seen.add(key)
        deduped.append(t)
    if len(deduped) < len(all_trades):
        print(f"\nDeduplication: removed {len(all_trades) - len(deduped)} duplicate 미래에셋 trades")
    all_trades = deduped

    print(f"\n{'='*60}\nTOTAL: {len(all_trades)} trades, {len(all_deposits)} deposits")

    filtered = filter_deposits(all_deposits)
    print(f"Filtered deposits: {len(filtered)}")

    computed = aggregate(all_trades)
    rate = compare(computed, GROUND_TRUTH)

    write_csv(all_trades, os.path.join(OUTPUT_DIR, 'trades_all.csv'))
    write_by_stock_xlsx(all_trades, os.path.join(OUTPUT_DIR, 'trades_by_stock.xlsx'))
    write_stock_md(all_trades, OUTPUT_DIR)
    write_deposits(filtered, os.path.join(OUTPUT_DIR, 'deposits.json'))
    write_summary(all_trades, filtered, os.path.join(OUTPUT_DIR, 'account', 'trades_summary.md'))
    write_account_md(filtered, OUTPUT_DIR)

    print(f"\n{'='*60}\nDONE. Match rate: {rate:.1f}%\n{'='*60}")


if __name__ == '__main__':
    main()
